"""
generate_report_summary.py (EZ_table0227)
========================================
Generate report_summary.xlsx with the SAME sheet layout / cell positions as
EZ_table0226/report_summary.xlsx (盈再表「美股」原始資料區版面).

This script intentionally mirrors the layout contract used by:
  EZ_table0226/generate_report_summary.py

Usage:
    python generate_report_summary.py
    python generate_report_summary.py AAPL
    python generate_report_summary.py 9022.T JPY
    python generate_report_summary.py AAPL "" "C:\\Temp\\report_temp_AAPL.xlsx"   # VBA override output
"""

from __future__ import annotations

import os
import sys
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook

# ========= Configuration =========
DEFAULT_TICKER = "CM"  # keep consistent with the existing EZ_table0227/report_summary.xlsx sample
DEFAULT_CURRENCY = None  # Auto-detect from financialCurrency; set e.g. "JPY" to override

HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(HERE, "report_summary.xlsx")
SHEET_NAME = "美股"

# CLI arg[3] can override the output path (used by VBA auto-fetch)
_output_override = sys.argv[3] if len(sys.argv) > 3 else None
OUTPUT_FILE_ACTUAL = _output_override if _output_override else OUTPUT_FILE


# =====================================================================
# Helper
# =====================================================================

def _sanitize(df: pd.DataFrame) -> pd.DataFrame:
    """Clean DataFrame for Excel output."""
    df = df.copy()
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.replace([np.inf, -np.inf], np.nan)
    return df


def _safe_val(val):
    """Return a safe value for writing to Excel (no NaN/inf, no formula injection)."""
    if val is None:
        return ""
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return ""
    # Prevent Excel formula injection from data values
    if isinstance(val, str) and len(val) > 0 and val[0] in ("=", "+", "-", "@"):
        return "'" + val
    return val


# =====================================================================
# Data Fetching Functions (yfinance)
# =====================================================================

def fetch_quarterly_income(ticker: yf.Ticker):
    """Fetch quarterly income statement: last 5 quarters + TTM (sum of last 4Q)."""
    df = ticker.quarterly_financials
    if df is None or df.empty:
        return pd.DataFrame(), pd.DataFrame()

    df.columns = pd.to_datetime(df.columns)
    df = df.sort_index(axis=1)

    last_q = df.iloc[:, -5:] if df.shape[1] >= 5 else df.copy()
    last_q.columns = last_q.columns.strftime("%Y-%m")

    if df.shape[1] >= 4:
        last_4q = df.iloc[:, -4:]
        ttm = last_4q.apply(pd.to_numeric, errors="coerce").sum(axis=1).to_frame(name="TTM")
    else:
        ttm = df.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
        ttm.columns = ["TTM"]

    combined = last_q.merge(ttm, left_index=True, right_index=True, how="left")

    keywords = ["unusual", "special", "restructuring", "non recurring"]
    unusual_mask = combined.index.str.lower().str.contains("|".join(keywords))
    unusual = combined.loc[unusual_mask] if unusual_mask.any() else pd.DataFrame()

    return _sanitize(combined), _sanitize(unusual) if not unusual.empty else unusual


def fetch_quarterly_balance_sheet(ticker: yf.Ticker) -> pd.DataFrame:
    """Fetch quarterly balance sheet: last 5 quarters + TTM (latest snapshot)."""
    df = ticker.quarterly_balance_sheet
    if df is None or df.empty:
        return pd.DataFrame()

    df.columns = pd.to_datetime(df.columns)
    df = df.sort_index(axis=1)

    last_q = df.iloc[:, -5:] if df.shape[1] >= 5 else df.copy()
    last_q.columns = last_q.columns.strftime("%Y-%m")

    ttm = df.iloc[:, -1:].copy()
    ttm.columns = ["TTM"]

    combined = last_q.merge(ttm, left_index=True, right_index=True, how="left")
    return _sanitize(combined)


def fetch_annual_income(ticker: yf.Ticker) -> pd.DataFrame:
    """Fetch annual income statement: last 5 years + TTM."""
    df = ticker.financials
    if df is None or df.empty:
        return pd.DataFrame()

    df.columns = pd.to_datetime(df.columns)
    df = df.sort_index(axis=1)

    current_year = datetime.today().year
    df = df.loc[:, df.columns.year < current_year]
    if df.empty:
        return pd.DataFrame()

    last_y = df.iloc[:, -5:] if df.shape[1] >= 5 else df.copy()
    last_y.columns = last_y.columns.strftime("%Y")

    ttm = df.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
    ttm.columns = ["TTM"]

    combined = last_y.merge(ttm, left_index=True, right_index=True, how="left")
    return _sanitize(combined)


def fetch_annual_balance_sheet(ticker: yf.Ticker) -> pd.DataFrame:
    """Fetch annual balance sheet: last 5 years + TTM (latest snapshot)."""
    df = ticker.balance_sheet
    if df is None or df.empty:
        return pd.DataFrame()

    df.columns = pd.to_datetime(df.columns)
    df = df.sort_index(axis=1)

    last_y = df.iloc[:, -5:] if df.shape[1] >= 5 else df.copy()
    last_y.columns = last_y.columns.strftime("%Y")

    ttm = df.iloc[:, -1:].copy()
    ttm.columns = ["TTM"]

    combined = last_y.merge(ttm, left_index=True, right_index=True, how="left")
    return _sanitize(combined)


def fetch_annual_cashflow(ticker: yf.Ticker) -> pd.DataFrame:
    """Fetch annual cash flow: last 5 years + TTM (4Q sum or fallback)."""
    cf_y = ticker.cashflow
    if cf_y is None or cf_y.empty:
        return pd.DataFrame()

    cf_y.columns = pd.to_datetime(cf_y.columns)
    cf_y = cf_y.sort_index(axis=1)

    last_y = cf_y.iloc[:, -5:] if cf_y.shape[1] >= 5 else cf_y.copy()
    last_y.columns = last_y.columns.strftime("%Y")

    try:
        cf_q = ticker.quarterly_cashflow
        if cf_q is not None and not cf_q.empty:
            cf_q.columns = pd.to_datetime(cf_q.columns)
            cf_q = cf_q.sort_index(axis=1)
            if cf_q.shape[1] >= 4:
                ttm = cf_q.iloc[:, -4:].apply(pd.to_numeric, errors="coerce").sum(axis=1).to_frame(name="TTM")
            else:
                ttm = cf_q.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
                ttm.columns = ["TTM"]
        else:
            ttm = cf_y.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
            ttm.columns = ["TTM"]
    except Exception:
        ttm = cf_y.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
        ttm.columns = ["TTM"]

    combined = last_y.merge(ttm, left_index=True, right_index=True, how="left")
    return _sanitize(combined)


def fetch_company_profile(ticker: yf.Ticker) -> dict:
    """Fetch company profile from ticker.info."""
    try:
        info = ticker.info
    except Exception:
        return {}
    if not info:
        return {}
    return {
        "longName": info.get("longName", ""),
        "symbol": info.get("symbol", ""),
        "exchange": info.get("exchange", ""),
        "sector": info.get("sector", ""),
        "industry": info.get("industry", ""),
        "country": info.get("country", ""),
        "currency": info.get("currency", ""),
        "financialCurrency": info.get("financialCurrency", "USD"),
        "currentPrice": info.get("currentPrice"),
        "marketCap": info.get("marketCap"),
        "enterpriseValue": info.get("enterpriseValue"),
        "trailingPE": info.get("trailingPE"),
        "forwardPE": info.get("forwardPE"),
        "trailingEps": info.get("trailingEps"),
        "dividendYield": info.get("dividendYield"),
        "payoutRatio": info.get("payoutRatio"),
        "beta": info.get("beta"),
        "fiftyTwoWeekHigh": info.get("fiftyTwoWeekHigh"),
        "fiftyTwoWeekLow": info.get("fiftyTwoWeekLow"),
        "longBusinessSummary": info.get("longBusinessSummary", ""),
        "website": info.get("website", ""),
        "totalRevenue": info.get("totalRevenue"),
        "netIncomeToCommon": info.get("netIncomeToCommon"),
    }


def fetch_share_capital(ticker: yf.Ticker) -> dict:
    """Fetch share capital data from ticker.info."""
    try:
        info = ticker.info
    except Exception:
        return {}
    if not info:
        return {}
    return {
        "Shares Outstanding": info.get("sharesOutstanding"),
        "Float Shares": info.get("floatShares"),
        "Implied Shares Outstanding": info.get("impliedSharesOutstanding"),
        "Held by Insiders (%)": info.get("heldPercentInsiders"),
        "Held by Institutions (%)": info.get("heldPercentInstitutions"),
        "Short Shares": info.get("sharesShort"),
        "Short Prior Month": info.get("sharesShortPriorMonth"),
        "Short % of Shares Outstanding": info.get("shortPercentOfFloat"),
    }


def fetch_eps_earnings(ticker: yf.Ticker) -> dict:
    """Fetch EPS / earnings data (net income annual/quarterly + trailing EPS)."""
    result = {"annual": {}, "quarterly": {}, "ttm_eps": None}

    # Annual net income from income statement
    try:
        is_y = ticker.income_stmt
        if is_y is not None and not is_y.empty:
            is_y.columns = pd.to_datetime(is_y.columns)
            is_y = is_y.sort_index(axis=1)
            for label in ["Net Income", "Net Income Common Stockholders"]:
                if label in is_y.index:
                    for col in is_y.columns:
                        val = is_y.loc[label, col]
                        if pd.notna(val):
                            result["annual"][col.strftime("%Y")] = val
                    break
    except Exception:
        pass

    # Quarterly net income
    try:
        is_q = ticker.quarterly_income_stmt
        if is_q is not None and not is_q.empty:
            is_q.columns = pd.to_datetime(is_q.columns)
            is_q = is_q.sort_index(axis=1)
            for label in ["Net Income", "Net Income Common Stockholders"]:
                if label in is_q.index:
                    for col in is_q.columns:
                        val = is_q.loc[label, col]
                        if pd.notna(val):
                            result["quarterly"][col.strftime("%Y-%m-%d")] = val
                    break
    except Exception:
        pass

    # TTM EPS from info
    try:
        info = ticker.info
        result["ttm_eps"] = info.get("trailingEps")
    except Exception:
        pass

    return result


def fetch_historical_prices_adj(ticker: yf.Ticker, start_date: str = "2014-01-01") -> dict:
    """Fetch monthly historical prices with Adj Close and corporate actions.
    Returns dict {'latest': {...}, 'rows': [...]} matching Historical_stock_price_adj(new).py."""
    empty = {"latest": None, "rows": []}
    try:
        end_date = (datetime.today() + timedelta(days=1)).strftime("%Y-%m-%d")
        df = ticker.history(start=start_date, end=end_date, auto_adjust=False)
    except Exception:
        return empty
    if df is None or df.empty:
        return empty

    # Latest trading day
    latest_row = df.iloc[-1]
    latest_date = df.index[-1].to_pydatetime().replace(tzinfo=None)
    latest = {
        "date": latest_date,
        "Open": latest_row.get("Open"),
        "High": latest_row.get("High"),
        "Low": latest_row.get("Low"),
        "Close": latest_row.get("Close"),
        "Adj Close": latest_row.get("Adj Close"),
        "Volume": int(latest_row.get("Volume")) if pd.notna(latest_row.get("Volume")) else None,
    }

    # Monthly resampling (month start), newest first
    monthly_df = (
        df.resample("MS")
        .agg({"Open": "first", "High": "max", "Low": "min",
              "Close": "last", "Adj Close": "last", "Volume": "sum"})
        .sort_index(ascending=False)
    )

    # Corporate Actions
    try:
        actions = ticker.actions
        if actions is not None and not actions.empty:
            actions = actions.loc[start_date:end_date]
            if "Dividends" not in actions.columns:
                actions["Dividends"] = 0
            if "Stock Splits" not in actions.columns:
                actions["Stock Splits"] = 0
            actions = actions[(actions["Dividends"] != 0) | (actions["Stock Splits"] != 0)]
        else:
            actions = pd.DataFrame()
    except Exception:
        actions = pd.DataFrame()

    # Build rows list (same structure as Historical_stock_price_adj(new).py)
    rows = []
    for dt, row in monthly_df.iterrows():
        rows.append({
            "date": dt.to_pydatetime().replace(tzinfo=None),
            "type": "monthly",
            "Open": row.get("Open"),
            "High": row.get("High"),
            "Low": row.get("Low"),
            "Close": row.get("Close"),
            "Adj Close": row.get("Adj Close"),
            "Volume": int(row.get("Volume")) if pd.notna(row.get("Volume")) else None,
        })

    if not actions.empty:
        for dt, row in actions.iterrows():
            d = dt.to_pydatetime().replace(tzinfo=None)
            div = row.get("Dividends", 0)
            spl = row.get("Stock Splits", 0)
            if pd.notna(div) and div != 0:
                rows.append({"date": d, "type": "dividend", "amount": div, "year": dt.year})
            if pd.notna(spl) and spl != 0:
                rows.append({"date": d, "type": "split", "amount": spl})

    rows.sort(key=lambda x: x["date"], reverse=True)
    return {"latest": latest, "rows": rows}


def fetch_exchange_rate(currency_code: str) -> dict:
    """Fetch USD/XXX exchange rate."""
    if not currency_code or currency_code.upper() == "USD":
        return {"rate": 1.0, "pair": "USD/USD"}
    symbol = f"USD{currency_code.upper()}=X"
    try:
        t = yf.Ticker(symbol)
        data = t.history(period="5d")
        if data is not None and not data.empty:
            latest = data.iloc[-1]
            return {
                "pair": f"USD/{currency_code.upper()}",
                "rate": latest["Close"],
                "open": latest.get("Open"),
                "high": latest.get("High"),
                "low": latest.get("Low"),
            }
    except Exception:
        pass
    return {"rate": 1.0, "pair": f"USD/{currency_code.upper()}"}


# =====================================================================
# Excel Writing Functions (layout contract)
# =====================================================================

def write_front_page(ws, profile, eps_data, share_cap, forex, net_income_ttm):
    """Write front-page cells: A1, A2, I1, F15, K3, Y16, Y23, Y24, W25, E9, K10-K12."""
    company_line = (
        f"{profile.get('longName', '')} {profile.get('symbol', '')} "
        f"({profile.get('exchange', '')}) : "
        f"{profile.get('sector', '')}*{profile.get('industry', '')}"
    )
    ws.cell(row=1, column=1, value=company_line)
    ws.cell(row=2, column=1, value=profile.get("symbol", ""))
    ws.cell(row=1, column=9, value=datetime.today().strftime("%Y-%m-%d"))
    ws.cell(row=15, column=6, value=_safe_val(forex.get("rate", 1.0)))
    ws.cell(row=3, column=11, value=_safe_val(profile.get("currentPrice")))
    ws.cell(row=16, column=25, value=_safe_val(profile.get("longBusinessSummary", "")))

    mc = profile.get("marketCap")
    ws.cell(row=23, column=25, value=round(mc / 1_000_000, 2) if mc is not None else "")
    ws.cell(row=24, column=25, value=_safe_val(profile.get("currentPrice")))
    ws.cell(row=25, column=23, value=_safe_val(profile.get("financialCurrency", "USD")))

    ws.cell(row=9, column=5, value=round(net_income_ttm / 1_000_000, 2) if net_income_ttm is not None else "")
    ws.cell(row=10, column=11, value=12)
    ws.cell(row=11, column=11, value=_safe_val(profile.get("payoutRatio")))
    ws.cell(row=12, column=11, value=0.25)


def write_financial_table(ws, df, start_col, section_title, section_num):
    """Write IS/BS/CFS with a shared contract: header row 1, columns row 2, data from row 3."""
    if df.empty:
        ws.cell(row=1, column=start_col, value=f"{section_num} / 9 {section_title}")
        return

    ws.cell(row=1, column=start_col, value=f"{section_num} / 9 {section_title}")
    ws.cell(row=2, column=start_col, value="Item")
    for ci, col_name in enumerate(df.columns, start=1):
        ws.cell(row=2, column=start_col + ci, value=col_name)

    for ri, (item_name, row_data) in enumerate(df.iterrows(), start=3):
        ws.cell(row=ri, column=start_col, value=item_name)
        for ci, col_name in enumerate(df.columns, start=1):
            ws.cell(row=ri, column=start_col + ci, value=_safe_val(row_data[col_name]))


def write_unusual_items(ws, unusual_df, start_col, data_row_count):
    """Append unusual items below the main quarterly IS table."""
    if unusual_df.empty:
        return
    sep_row = 3 + data_row_count + 1
    ws.cell(row=sep_row, column=start_col, value="[Unusual / Special Items]")
    for ri, (item_name, row_data) in enumerate(unusual_df.iterrows(), start=sep_row + 1):
        ws.cell(row=ri, column=start_col, value=item_name)
        for ci, col_name in enumerate(unusual_df.columns, start=1):
            ws.cell(row=ri, column=start_col + ci, value=_safe_val(row_data[col_name]))


def write_price_history(ws, price_data, start_col):
    """Write historical price section matching Historical_stock_price_adj(new).py layout.
    start_col   (BN=66): section title + year for dividend rows
    start_col+1 (BO=67): Date header and dates
    start_col+2..+7 (BP..BU): Open / High / Low / Close / Adj Close / Volume
    Row 1: section title (BN) + column headers (BO:BU)
    Row 2-3: empty (BN2:BU3 reserved, matching 盈再表 layout)
    Row 4: latest trading day
    Row 5+: monthly rows + corporate action rows (newest first)
    """
    DATE_FMT = "yyyy/m/d;@"
    ws.cell(row=1, column=start_col, value="5 / 9 Historical stock price (adj)")

    if not price_data or (not price_data.get("rows") and price_data.get("latest") is None):
        return

    # Row 1: column headers at BO..BU (start_col+1..+7)
    for i, h in enumerate(["Date", "Open", "High", "Low", "Close", "Adj Close", "Volume"]):
        ws.cell(row=1, column=start_col + 1 + i, value=h)

    # Row 2-3: empty (BN2:BU3 reserved)

    # Row 4: latest trading day
    latest = price_data.get("latest")
    if latest:
        ws.cell(row=4, column=start_col + 1, value=latest["date"]).number_format = DATE_FMT
        ws.cell(row=4, column=start_col + 2, value=_safe_val(latest["Open"]))
        ws.cell(row=4, column=start_col + 3, value=_safe_val(latest["High"]))
        ws.cell(row=4, column=start_col + 4, value=_safe_val(latest["Low"]))
        ws.cell(row=4, column=start_col + 5, value=_safe_val(latest["Close"]))
        ws.cell(row=4, column=start_col + 6, value=_safe_val(latest["Adj Close"]))
        ws.cell(row=4, column=start_col + 7, value=latest["Volume"])

    # Row 5+: monthly rows + corporate action rows
    for i, r in enumerate(price_data.get("rows", [])):
        excel_row = 5 + i
        if r["type"] == "monthly":
            ws.cell(row=excel_row, column=start_col + 1, value=r["date"]).number_format = DATE_FMT
            ws.cell(row=excel_row, column=start_col + 2, value=_safe_val(r["Open"]))
            ws.cell(row=excel_row, column=start_col + 3, value=_safe_val(r["High"]))
            ws.cell(row=excel_row, column=start_col + 4, value=_safe_val(r["Low"]))
            ws.cell(row=excel_row, column=start_col + 5, value=_safe_val(r["Close"]))
            ws.cell(row=excel_row, column=start_col + 6, value=_safe_val(r["Adj Close"]))
            ws.cell(row=excel_row, column=start_col + 7, value=r["Volume"])
        elif r["type"] == "dividend":
            ws.cell(row=excel_row, column=start_col, value=r["year"])
            ws.cell(row=excel_row, column=start_col + 1, value=r["date"]).number_format = DATE_FMT
            ws.cell(row=excel_row, column=start_col + 2, value=_safe_val(r["amount"]))
            ws.cell(row=excel_row, column=start_col + 3, value="Dividend")
        elif r["type"] == "split":
            ws.cell(row=excel_row, column=start_col + 1, value=r["date"]).number_format = DATE_FMT
            ws.cell(row=excel_row, column=start_col + 2, value=_safe_val(r["amount"]))
            ws.cell(row=excel_row, column=start_col + 3, value="Split")


def write_profile_section(ws, profile, start_col):
    ws.cell(row=1, column=start_col, value="6 / 9 Company Profile")
    items = [
        ("Company Name", profile.get("longName")),
        ("Ticker", profile.get("symbol")),
        ("Exchange", profile.get("exchange")),
        ("Sector", profile.get("sector")),
        ("Industry", profile.get("industry")),
        ("Country", profile.get("country")),
        ("Currency", profile.get("currency")),
        ("Financial Currency", profile.get("financialCurrency")),
        ("Current Price", profile.get("currentPrice")),
        ("Market Cap", profile.get("marketCap")),
        ("Enterprise Value", profile.get("enterpriseValue")),
        ("Trailing PE", profile.get("trailingPE")),
        ("Forward PE", profile.get("forwardPE")),
        ("EPS (TTM)", profile.get("trailingEps")),
        ("Dividend Yield", profile.get("dividendYield")),
        ("Payout Ratio", profile.get("payoutRatio")),
        ("Beta", profile.get("beta")),
        ("52 Week High", profile.get("fiftyTwoWeekHigh")),
        ("52 Week Low", profile.get("fiftyTwoWeekLow")),
        ("Revenue (TTM)", profile.get("totalRevenue")),
        ("Net Income (TTM)", profile.get("netIncomeToCommon")),
        ("Website", profile.get("website")),
        ("Description", profile.get("longBusinessSummary")),
    ]
    for ri, (item, value) in enumerate(items, start=2):
        ws.cell(row=ri, column=start_col, value=item)
        ws.cell(row=ri, column=start_col + 1, value=_safe_val(value))


def write_forex_section(ws, forex, start_col):
    ws.cell(row=1, column=start_col, value="7 / 9 Foreign exchange rate")
    items = [
        ("Currency Pair", forex.get("pair", "")),
        ("Exchange Rate (Close)", forex.get("rate")),
        ("Open", forex.get("open")),
        ("High", forex.get("high")),
        ("Low", forex.get("low")),
    ]
    for ri, (item, value) in enumerate(items, start=2):
        ws.cell(row=ri, column=start_col, value=item)
        ws.cell(row=ri, column=start_col + 1, value=_safe_val(value))


def write_market_cap_section(ws, profile, share_cap, eps_data, start_col):
    ws.cell(row=1, column=start_col, value="9 / 9 Market capitalization")
    mc = profile.get("marketCap")
    ev = profile.get("enterpriseValue")
    items = [
        ("Market Cap ($M)", round(mc / 1_000_000, 2) if mc else ""),
        ("Enterprise Value ($M)", round(ev / 1_000_000, 2) if ev else ""),
        ("Trailing PE", profile.get("trailingPE")),
        ("Forward PE", profile.get("forwardPE")),
        ("EPS (TTM)", profile.get("trailingEps")),
        ("Dividend Yield", profile.get("dividendYield")),
        ("Payout Ratio", profile.get("payoutRatio")),
        ("Beta", profile.get("beta")),
        ("52 Week High", profile.get("fiftyTwoWeekHigh")),
        ("52 Week Low", profile.get("fiftyTwoWeekLow")),
        ("Revenue (TTM)", profile.get("totalRevenue")),
        ("Net Income (TTM)", profile.get("netIncomeToCommon")),
        ("Shares Outstanding", share_cap.get("Shares Outstanding")),
        ("Float Shares", share_cap.get("Float Shares")),
        ("Implied Shares Outstanding", share_cap.get("Implied Shares Outstanding")),
        ("Held by Insiders (%)", share_cap.get("Held by Insiders (%)")),
        ("Held by Institutions (%)", share_cap.get("Held by Institutions (%)")),
        ("Short Shares", share_cap.get("Short Shares")),
        ("Short Prior Month", share_cap.get("Short Prior Month")),
        ("Short % of Shares Outstanding", share_cap.get("Short % of Shares Outstanding")),
    ]

    if eps_data.get("ttm_eps") is not None:
        items.append(("TTM EPS", eps_data["ttm_eps"]))
    if eps_data.get("annual"):
        items.append(("[Annual Earnings]", ""))
        for year, val in eps_data["annual"].items():
            items.append((f"Earnings {year}", val))
    if eps_data.get("quarterly"):
        items.append(("[Quarterly Earnings]", ""))
        for date_str, val in eps_data["quarterly"].items():
            items.append((f"Earnings {date_str}", val))

    for ri, (item, value) in enumerate(items, start=2):
        ws.cell(row=ri, column=start_col, value=item)
        ws.cell(row=ri, column=start_col + 1, value=_safe_val(value))


# =====================================================================
# Main
# =====================================================================

def main():
    ticker_symbol = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_TICKER
    currency_override = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_CURRENCY

    print("=" * 60)
    print("  Report Summary Generator (EZ_table0227)")
    print(f"  Ticker: {ticker_symbol}")
    print("=" * 60)

    ticker = yf.Ticker(ticker_symbol)

    print("1/10 Fetching quarterly income statement...")
    is_q, unusual_q = fetch_quarterly_income(ticker)

    print("2/10 Fetching quarterly balance sheet...")
    bs_q = fetch_quarterly_balance_sheet(ticker)

    print("3/10 Fetching annual income statement...")
    is_y = fetch_annual_income(ticker)

    print("4/10 Fetching annual balance sheet...")
    bs_y = fetch_annual_balance_sheet(ticker)

    print("5/10 Fetching annual cash flow statement...")
    cfs_y = fetch_annual_cashflow(ticker)

    print("6/10 Fetching company profile...")
    profile = fetch_company_profile(ticker)

    print("7/10 Fetching share capital data...")
    share_cap = fetch_share_capital(ticker)

    print("8/10 Fetching EPS / earnings data...")
    eps_data = fetch_eps_earnings(ticker)

    print("9/10 Fetching historical prices (adjusted) & actions...")
    combined_prices = fetch_historical_prices_adj(ticker)

    fin_currency = currency_override or profile.get("financialCurrency", "USD")
    print(f"10/10 Fetching exchange rate (USD/{fin_currency})...")
    forex = fetch_exchange_rate(fin_currency)

    # Net Income TTM for front page
    net_income_ttm = None
    if not is_q.empty:
        for label in ["Net Income", "Net Income Common Stockholders"]:
            if label in is_q.index and "TTM" in is_q.columns:
                val = is_q.loc[label, "TTM"]
                if pd.notna(val):
                    net_income_ttm = val
                    break
    if net_income_ttm is None and profile.get("netIncomeToCommon") is not None:
        net_income_ttm = profile["netIncomeToCommon"]

    # Column positions (must match EZ_table0226 layout)
    COL_IS_Q = 31   # AE
    COL_BS_Q = 40   # AN
    COL_IS_Y = 49   # AW
    COL_BS_Y = 58   # BF
    COL_PRICE = 66  # BN
    COL_PROF = 76   # BX
    COL_FX = 79     # CA
    COL_CFS = 83    # CE
    COL_MCAP = 92   # CN

    print("\nWriting Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    write_front_page(ws, profile, eps_data, share_cap, forex, net_income_ttm)

    write_financial_table(ws, is_q, COL_IS_Q, "Quarterly Income Statement", 1)
    if not unusual_q.empty:
        write_unusual_items(ws, unusual_q, COL_IS_Q, len(is_q))

    write_financial_table(ws, bs_q, COL_BS_Q, "Quarterly Balance Sheet", 2)
    write_financial_table(ws, is_y, COL_IS_Y, "Annual Income Statement", 3)
    write_financial_table(ws, bs_y, COL_BS_Y, "Annual Balance Sheet", 4)
    write_price_history(ws, combined_prices, COL_PRICE)
    write_profile_section(ws, profile, COL_PROF)
    write_forex_section(ws, forex, COL_FX)
    write_financial_table(ws, cfs_y, COL_CFS, "Annual Cash Flow", 8)
    write_market_cap_section(ws, profile, share_cap, eps_data, COL_MCAP)

    wb.save(OUTPUT_FILE_ACTUAL)
    print(f"\n[OK] Saved: {OUTPUT_FILE_ACTUAL}")


if __name__ == "__main__":
    main()

