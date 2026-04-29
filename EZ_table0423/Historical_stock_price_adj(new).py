import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
import os


# ========= 1️⃣ 單一股票輸入 =========
#ticker = "AAPL"   # 🔥 在這裡改股票代號
#ticker = "DIOD"   # 🔥 在這裡改股票代號
#ticker = "MITSY"   # 🔥 在這裡改股票代號
ticker = "6415.TW"   # 🔥 在這裡改股票代號



start_date = "2014-01-01"
end_date = (datetime.today() + timedelta(days=1)).strftime("%Y-%m-%d")


print(f"正在處理 {ticker}...")

stock = yf.Ticker(ticker)

# ========= 2️⃣ 抓日線資料 =========
df = stock.history(start=start_date, end=end_date, auto_adjust=False)

if df.empty:
    raise ValueError("⚠ 無資料")

# ========= 3️⃣ 最近交易日資料 =========
latest = df.iloc[-1]
latest_date = df.index[-1].to_pydatetime().replace(tzinfo=None)

# ========= 4️⃣ 每月1日基準 =========
monthly_df = df.resample('MS').agg({
    'Open': 'first',
    'High': 'max',
    'Low': 'min',
    'Close': 'last',
    'Adj Close': 'last',
    'Volume': 'sum'
}).sort_index(ascending=False)

# ========= 5️⃣ 取得 Corporate Actions =========
actions = stock.actions
if not actions.empty:
    actions = actions.loc[start_date:end_date]
    if 'Dividends' not in actions.columns:
        actions['Dividends'] = 0
    if 'Stock Splits' not in actions.columns:
        actions['Stock Splits'] = 0
    actions = actions[
        (actions['Dividends'] != 0) |
        (actions['Stock Splits'] != 0)
    ]
else:
    actions = pd.DataFrame()

# ========= 6️⃣ 建立所有資料列 =========
rows = []

for date, row in monthly_df.iterrows():
    rows.append({
        'date': date.to_pydatetime().replace(tzinfo=None),
        'type': 'monthly',
        'Open': row['Open'],
        'High': row['High'],
        'Low': row['Low'],
        'Close': row['Close'],
        'Adj Close': row['Adj Close'],
        'Volume': int(row['Volume']),
    })

if not actions.empty:
    for date, row in actions.iterrows():
        d = date.to_pydatetime().replace(tzinfo=None)
        div = row['Dividends']
        spl = row['Stock Splits']
        if div != 0:
            rows.append({
                'date': d,
                'type': 'dividend',
                'amount': div,
                'year': date.year,
            })
        if spl != 0:
            rows.append({
                'date': d,
                'type': 'split',
                'amount': spl,
            })

# 依日期降序排列（穩定排序：同日期中 dividend 先於 split）
rows.sort(key=lambda x: x['date'], reverse=True)

# ========= 7️⃣ 輸出 Excel =========
wb = Workbook()
ws = wb.active

DATE_FMT = 'yyyy/m/d;@'
COL_WIDTH = 13.0

for col_letter in 'ABCDEFGH':
    ws.column_dimensions[col_letter].width = COL_WIDTH

# 第一列：標題
headers = ['Date', 'Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume']
for i, h in enumerate(headers):
    ws.cell(row=1, column=2 + i, value=h)
ws.cell(row=1, column=2).number_format = DATE_FMT

# 第二列：最近交易日
ws.cell(row=2, column=2, value=latest_date).number_format = DATE_FMT
ws.cell(row=2, column=3, value=latest['Open'])
ws.cell(row=2, column=4, value=latest['High'])
ws.cell(row=2, column=5, value=latest['Low'])
ws.cell(row=2, column=6, value=latest['Close'])
ws.cell(row=2, column=7, value=latest['Adj Close'])
ws.cell(row=2, column=8, value=int(latest['Volume']))

# 第三列起：月線 + 公司行動
for i, r in enumerate(rows):
    excel_row = 3 + i
    if r['type'] == 'monthly':
        ws.cell(row=excel_row, column=2, value=r['date']).number_format = DATE_FMT
        ws.cell(row=excel_row, column=3, value=r['Open'])
        ws.cell(row=excel_row, column=4, value=r['High'])
        ws.cell(row=excel_row, column=5, value=r['Low'])
        ws.cell(row=excel_row, column=6, value=r['Close'])
        ws.cell(row=excel_row, column=7, value=r['Adj Close'])
        ws.cell(row=excel_row, column=8, value=r['Volume'])
    elif r['type'] == 'dividend':
        ws.cell(row=excel_row, column=1, value=r['year'])
        ws.cell(row=excel_row, column=2, value=r['date']).number_format = DATE_FMT
        ws.cell(row=excel_row, column=3, value=r['amount'])
        ws.cell(row=excel_row, column=4, value='Dividend')
    elif r['type'] == 'split':
        ws.cell(row=excel_row, column=2, value=r['date']).number_format = DATE_FMT
        ws.cell(row=excel_row, column=3, value=r['amount'])
        ws.cell(row=excel_row, column=4, value='Split')

output_file = f"{ticker}_Historical_stock_price_adj(new).xlsx"
wb.save(output_file)

print(f"[OK] 已輸出：{os.path.abspath(output_file)}")
print("[DONE] 任務完成！")
