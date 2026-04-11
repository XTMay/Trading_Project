"""
stock_fetcher.py - yfinance stock data fetcher
===============================================
Fetches stock data and writes to a temp Excel file for VBA to import.

Usage:
  python stock_fetcher.py AAPL                    # write to default temp file
  python stock_fetcher.py AAPL /path/to/out.xlsx  # write to specified file
"""

import sys
import yfinance as yf
import numpy as np
import pandas as pd
import openpyxl

import tempfile, os
DEFAULT_OUTPUT = os.path.join(tempfile.gettempdir(), "stock_temp.xlsx")


# ============================================================
# Utilities
# ============================================================

def _clean(value):
    """Convert numpy/pandas types to native Python types."""
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        v = float(value)
        return "" if np.isnan(v) else round(v, 4)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    return value


def _write_df(ws, df, start_row=1, start_col=1, write_index=True, index_label=""):
    """Write a DataFrame to an openpyxl worksheet."""
    r, c = start_row, start_col

    if write_index:
        ws.cell(row=r, column=c, value=index_label or "")
        for j, col_name in enumerate(df.columns):
            ws.cell(row=r, column=c + 1 + j, value=str(col_name))
    else:
        for j, col_name in enumerate(df.columns):
            ws.cell(row=r, column=c + j, value=str(col_name))

    for i in range(len(df)):
        r += 1
        offset = 0
        if write_index:
            ws.cell(row=r, column=c, value=str(df.index[i]))
            offset = 1
        for j in range(len(df.columns)):
            ws.cell(row=r, column=c + offset + j, value=_clean(df.iloc[i, j]))


# ============================================================
# Data fetch functions
# ============================================================

def fetch_info(ticker, ws):
    """Stock overview."""
    info = ticker.info
    fields = [
        ("Symbol",          info.get("symbol", "")),
        ("Company",         info.get("shortName", "")),
        ("Industry",        info.get("industry", "")),
        ("Sector",          info.get("sector", "")),
        ("Country",         info.get("country", "")),
        ("Market Cap",      info.get("marketCap", "")),
        ("Enterprise Value",info.get("enterpriseValue", "")),
        ("PE (TTM)",        info.get("trailingPE", "")),
        ("PE (Forward)",    info.get("forwardPE", "")),
        ("PB",              info.get("priceToBook", "")),
        ("EPS (TTM)",       info.get("trailingEps", "")),
        ("Dividend Yield",  info.get("dividendYield", "")),
        ("Dividend Rate",   info.get("dividendRate", "")),
        ("52W High",        info.get("fiftyTwoWeekHigh", "")),
        ("52W Low",         info.get("fiftyTwoWeekLow", "")),
        ("Beta",            info.get("beta", "")),
        ("Avg Volume",      info.get("averageVolume", "")),
        ("Current Price",   info.get("currentPrice", "")),
    ]
    ws.cell(row=1, column=1, value="Field")
    ws.cell(row=1, column=2, value="Value")
    for i, (label, val) in enumerate(fields):
        ws.cell(row=i + 2, column=1, value=label)
        ws.cell(row=i + 2, column=2, value=_clean(val))


def fetch_history(ticker, ws, period="3mo"):
    """Price history."""
    hist = ticker.history(period=period)
    if hist.empty:
        ws.cell(row=1, column=1, value="No data")
        return
    hist.index = hist.index.strftime("%Y-%m-%d")
    _write_df(ws, hist, write_index=True, index_label="Date")


def fetch_financials(ticker, ws, report_type="income"):
    """Financial statements."""
    report_map = {
        "income":   ticker.financials,
        "balance":  ticker.balance_sheet,
        "cashflow": ticker.cashflow,
    }
    df = report_map.get(report_type)
    if df is None or df.empty:
        ws.cell(row=1, column=1, value="No data")
        return
    df.columns = [c.strftime("%Y-%m-%d") if hasattr(c, "strftime") else str(c)
                  for c in df.columns]
    _write_df(ws, df, write_index=True, index_label="Item")


def fetch_dividends(ticker, ws):
    """Dividends + stock splits."""
    divs = ticker.dividends.tail(20)
    ws.cell(row=1, column=1, value="Date")
    ws.cell(row=1, column=2, value="Dividend")
    for i, (date, val) in enumerate(divs.items()):
        ws.cell(row=i + 2, column=1, value=str(date)[:10])
        ws.cell(row=i + 2, column=2, value=round(float(val), 4))

    gap = 2 + len(divs) + 2
    splits = ticker.splits
    ws.cell(row=gap, column=1, value="Date")
    ws.cell(row=gap, column=2, value="Split Ratio")
    for i, (date, val) in enumerate(splits.items()):
        ws.cell(row=gap + 1 + i, column=1, value=str(date)[:10])
        ws.cell(row=gap + 1 + i, column=2, value=float(val))


def fetch_holders(ticker, ws):
    """Major + institutional holders."""
    r = 1
    mh = ticker.major_holders
    if mh is not None and not mh.empty:
        ws.cell(row=r, column=1, value="Major Holders")
        r += 1
        for i in range(len(mh)):
            ws.cell(row=r + i, column=1, value=_clean(mh.iloc[i, 0]))
            if len(mh.columns) > 1:
                ws.cell(row=r + i, column=2, value=_clean(mh.iloc[i, 1]))
        r += len(mh) + 1

    ih = ticker.institutional_holders
    if ih is not None and not ih.empty:
        ws.cell(row=r, column=1, value="Institutional Holders (Top 10)")
        r += 1
        top10 = ih.head(10)
        for j, c in enumerate(top10.columns):
            ws.cell(row=r, column=j + 1, value=str(c))
        for i in range(len(top10)):
            for j in range(len(top10.columns)):
                ws.cell(row=r + 1 + i, column=j + 1, value=_clean(top10.iloc[i, j]))


def fetch_recommendations(ticker, ws):
    """Analyst recommendations."""
    rec = ticker.recommendations
    if rec is None or rec.empty:
        ws.cell(row=1, column=1, value="No data")
        return
    _write_df(ws, rec, write_index=False)


# ============================================================
# Main
# ============================================================

def main():
    if len(sys.argv) < 2:
        print("Usage: python stock_fetcher.py <SYMBOL> [output.xlsx]")
        sys.exit(1)

    symbol = sys.argv[1].strip().upper()
    output_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT

    print(f"Fetching {symbol}...")

    ticker = yf.Ticker(symbol)
    info = ticker.info
    if not info.get("shortName"):
        print(f"Error: invalid symbol {symbol}")
        sys.exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_tasks = [
        ("Info",          lambda ws: fetch_info(ticker, ws)),
        ("History_",      lambda ws: fetch_history(ticker, ws)),
        ("Income",        lambda ws: fetch_financials(ticker, ws, "income")),
        ("Balance Sheet", lambda ws: fetch_financials(ticker, ws, "balance")),
        ("Cash Flow",     lambda ws: fetch_financials(ticker, ws, "cashflow")),
        ("Dividends",     lambda ws: fetch_dividends(ticker, ws)),
        ("Holders",       lambda ws: fetch_holders(ticker, ws)),
        ("Recommend",     lambda ws: fetch_recommendations(ticker, ws)),
    ]

    for name, fetch_func in sheet_tasks:
        try:
            print(f"  [{name}]...")
            ws = wb.create_sheet(name)
            fetch_func(ws)
        except Exception as e:
            print(f"  [WARN] {name}: {e}")
            if name not in wb.sheetnames:
                ws = wb.create_sheet(name)
            wb[name].cell(row=1, column=1, value=f"Failed: {e}")

    wb.save(output_path)
    print(f"[OK] {symbol} ({info.get('shortName', '')}) -> {output_path}")


if __name__ == "__main__":
    main()
