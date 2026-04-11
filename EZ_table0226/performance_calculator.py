"""
performance_calculator.py
XIRR & Buffett Distance Engine for 盈再表 portfolio performance analysis.

Implements VBA Module8 calculations:
  - XIRR (Extended Internal Rate of Return)
  - Annual Return (CAGR)
  - Buffett Distance Score
  - Holding Ratio
  - Portfolio Performance Evaluation

Usage:
    python performance_calculator.py                    # demo with sample data
    python performance_calculator.py portfolio.xlsx     # evaluate from Excel file
"""

import sys
import math
import pandas as pd
import numpy as np
from datetime import datetime, date
from openpyxl import Workbook

from generate_report_summary import _safe_val

# =====================================================================
# XIRR Calculation
# =====================================================================


def _xirr_npv(rate, cashflows, dates):
    """Calculate NPV for a given rate using XIRR day-counting.

    NPV = Σ cashflow_i / (1 + rate) ^ ((date_i - date_0) / 365)
    """
    if rate <= -1.0:
        return float("inf")
    d0 = dates[0]
    npv = 0.0
    for cf, d in zip(cashflows, dates):
        days = (d - d0).days
        npv += cf / (1.0 + rate) ** (days / 365.0)
    return npv


def calc_xirr(cashflows, dates, guess=0.1):
    """Calculate XIRR using scipy.optimize.brentq (Brent's method).

    Solves for r: Σ cf_i / (1 + r)^((d_i - d_0)/365) = 0

    Args:
        cashflows: list of float, cash flow amounts (negative = outflow, positive = inflow).
        dates: list of datetime.date or datetime.datetime, corresponding dates.
        guess: initial guess (unused with brentq, kept for API compat).

    Returns:
        float: XIRR rate (e.g. 0.10 = 10%), or None if solver fails.
    """
    if not cashflows or not dates or len(cashflows) != len(dates):
        return None
    if len(cashflows) < 2:
        return None

    # Ensure dates are date objects
    dates = [d.date() if isinstance(d, datetime) else d for d in dates]

    # Sort by date
    paired = sorted(zip(dates, cashflows), key=lambda x: x[0])
    dates = [p[0] for p in paired]
    cashflows = [p[1] for p in paired]

    # Check that we have both positive and negative cash flows
    has_pos = any(cf > 0 for cf in cashflows)
    has_neg = any(cf < 0 for cf in cashflows)
    if not (has_pos and has_neg):
        return None

    try:
        from scipy.optimize import brentq
        result = brentq(
            lambda r: _xirr_npv(r, cashflows, dates),
            -0.99, 10.0,
            xtol=1e-8,
            maxiter=1000,
        )
        return result
    except Exception:
        # Fallback: Newton's method
        return _xirr_newton(cashflows, dates, guess)


def _xirr_newton(cashflows, dates, guess=0.1, tol=1e-8, max_iter=500):
    """Newton's method fallback for XIRR."""
    d0 = dates[0]
    rate = guess
    for _ in range(max_iter):
        npv = 0.0
        dnpv = 0.0
        for cf, d in zip(cashflows, dates):
            t = (d - d0).days / 365.0
            denom = (1.0 + rate) ** t
            if denom == 0:
                return None
            npv += cf / denom
            if t != 0:
                dnpv -= t * cf / ((1.0 + rate) ** (t + 1))
        if abs(dnpv) < 1e-14:
            return None
        new_rate = rate - npv / dnpv
        if abs(new_rate - rate) < tol:
            return new_rate
        rate = new_rate
    return None


# =====================================================================
# Annual Return (CAGR)
# =====================================================================


def calc_annual_return(current_value, initial_value, years):
    """Calculate annualized return (CAGR).

    Formula: (current / initial) ^ (1/years) - 1

    Args:
        current_value: Current portfolio/stock value.
        initial_value: Initial investment value.
        years: Number of years held.

    Returns:
        float: Annual return rate, or None.
    """
    if (current_value is None or initial_value is None or years is None
            or initial_value <= 0 or years <= 0 or current_value <= 0):
        return None
    return (current_value / initial_value) ** (1.0 / years) - 1


# =====================================================================
# Buffett Distance Score (from VBA Module8 pf(f) function)
# =====================================================================

# Constants from VBA
_BUFFETT_YEARS = 50          # Buffett's track record length
_BUFFETT_MULTIPLIER = 1.2    # 20% annual return base
_BUFFETT_CUMULATIVE = 9100   # ≈ 1.2^50
_BASELINE_YEARS = 8          # Reference baseline period
_BASELINE_RETURN = 0.12      # 12% reference annual return
# Normalization: (1.2^(50-8))^0.5 * log(9100)/log(1.12)
_NORMALIZATION = (
    (_BUFFETT_MULTIPLIER ** (_BUFFETT_YEARS - _BASELINE_YEARS)) ** 0.5
    * math.log(_BUFFETT_CUMULATIVE) / math.log(1 + _BASELINE_RETURN)
)


def calc_buffett_distance(years_held, annual_return):
    """Calculate Buffett Distance Score.

    From VBA Module8:
      A = (1.2 ^ (50 - years_held)) ^ 0.5
      B = log(9100) / log(1 + annual_return)
      Score = 100 × A × B / normalization_factor

    Where normalization = (1.2^42)^0.5 × log(9100)/log(1.12) ≈ 3701

    Args:
        years_held: Number of years the investment has been held.
        annual_return: Annualized return rate (e.g. 0.12 = 12%).

    Returns:
        dict with keys: score, rating, description.
    """
    if annual_return is None or annual_return <= 0 or years_held is None or years_held <= 0:
        return {"score": None, "rating": "N/A", "description": "Insufficient data or loss"}

    a = (_BUFFETT_MULTIPLIER ** (_BUFFETT_YEARS - years_held)) ** 0.5
    b = math.log(_BUFFETT_CUMULATIVE) / math.log(1 + annual_return)
    score = int(100 * a * b / _NORMALIZATION)

    if score <= 21:
        rating = "Buffett-level"
        description = "巴菲特等级！(Comparable to Buffett!)"
    elif score <= 100:
        rating = "Excellent"
        description = "非常优秀！(Excellent!)"
    elif score <= 168:
        rating = "Good"
        description = "加油！(Go for it!)"
    else:
        rating = "Needs review"
        description = "常回来看看讲义 (Review lecture transcript)"

    return {"score": score, "rating": rating, "description": description}


# =====================================================================
# Holding Ratio
# =====================================================================


def calc_holding_ratio(stock_value, cash_value):
    """Calculate stock holding ratio.

    Formula: stock_value / (stock_value + cash_value)

    Args:
        stock_value: Total market value of stock holdings.
        cash_value: Cash reserve value.

    Returns:
        float: Holding ratio (0 to 1), or None.
    """
    total = (stock_value or 0) + (cash_value or 0)
    if total <= 0:
        return None
    return (stock_value or 0) / total


# =====================================================================
# Portfolio Performance Evaluation
# =====================================================================


def evaluate_portfolio_performance(transactions_df):
    """Parse transaction list and compute XIRR + Buffett distance.

    Args:
        transactions_df: DataFrame with columns:
            - date: transaction date
            - amount: cash flow (negative=buy, positive=sell/dividend)
            - Optional: description

    Returns:
        dict with performance metrics.
    """
    if transactions_df is None or transactions_df.empty:
        return {"xirr": None, "annual_return": None, "buffett_distance": None}

    df = transactions_df.copy()

    # Ensure date column is datetime
    if "date" not in df.columns or "amount" not in df.columns:
        return {"xirr": None, "annual_return": None, "buffett_distance": None}

    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date")

    cashflows = df["amount"].tolist()
    dates = df["date"].tolist()

    # Calculate XIRR
    xirr = calc_xirr(cashflows, dates)

    # Calculate years held
    first_date = dates[0]
    last_date = dates[-1]
    years = (last_date - first_date).days / 365.0

    # Calculate total invested and current value for CAGR
    invested = sum(cf for cf in cashflows if cf < 0)
    received = sum(cf for cf in cashflows if cf > 0)
    cagr = calc_annual_return(abs(received), abs(invested), years) if years > 0 else None

    # Calculate Buffett distance
    buffett = calc_buffett_distance(years, xirr) if xirr else {
        "score": None, "rating": "N/A", "description": ""
    }

    # Holding ratio (assuming last positive CF is current value)
    total_invested = abs(invested)
    total_received = abs(received)

    return {
        "xirr": xirr,
        "annual_return": cagr,
        "years_held": round(years, 2),
        "total_invested": total_invested,
        "total_received": total_received,
        "buffett_distance": buffett,
    }


def evaluate_from_excel(excel_path, sheet_name=None):
    """Load transactions from Excel and evaluate performance.

    Expected Excel columns: date, amount (negative=buy, positive=sell/dividend).

    Args:
        excel_path: Path to Excel file with transactions.
        sheet_name: Optional sheet name.

    Returns:
        dict with performance metrics.
    """
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"Error reading {excel_path}: {e}")
        return None

    # Try to find date and amount columns
    date_col = None
    amount_col = None
    for col in df.columns:
        col_lower = str(col).lower()
        if "date" in col_lower or "日期" in col_lower:
            date_col = col
        if "amount" in col_lower or "金额" in col_lower or "cash" in col_lower:
            amount_col = col

    if date_col is None or amount_col is None:
        print("Could not find 'date' and 'amount' columns in the Excel file.")
        print(f"Available columns: {list(df.columns)}")
        return None

    df = df.rename(columns={date_col: "date", amount_col: "amount"})
    return evaluate_portfolio_performance(df)


def export_performance_to_excel(result, output_path="performance_report.xlsx"):
    """Write performance results to Excel.

    Args:
        result: dict from evaluate_portfolio_performance().
        output_path: Output file path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance"

    ws.cell(row=1, column=1, value="Portfolio Performance Report")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.today().strftime('%Y-%m-%d')}")

    row = 4
    items = [
        ("XIRR (年化报酬率)", result.get("xirr")),
        ("CAGR (年化成长率)", result.get("annual_return")),
        ("Years Held (持有年数)", result.get("years_held")),
        ("Total Invested (总投入)", result.get("total_invested")),
        ("Total Received (总回收)", result.get("total_received")),
    ]

    for label, val in items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=_safe_val(val))
        row += 1

    # Buffett Distance
    bd = result.get("buffett_distance", {})
    row += 1
    ws.cell(row=row, column=1, value="Buffett Distance Score")
    ws.cell(row=row, column=2, value=_safe_val(bd.get("score")))
    row += 1
    ws.cell(row=row, column=1, value="Rating")
    ws.cell(row=row, column=2, value=bd.get("rating", ""))
    row += 1
    ws.cell(row=row, column=1, value="Description")
    ws.cell(row=row, column=2, value=bd.get("description", ""))

    wb.save(output_path)
    print(f"Saved: {output_path}")


# =====================================================================
# CLI Entry Point
# =====================================================================

def main():
    if len(sys.argv) > 1:
        # Evaluate from Excel file
        excel_path = sys.argv[1]
        sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
        result = evaluate_from_excel(excel_path, sheet_name)
        if result:
            print(f"\n{'='*50}")
            print("  Portfolio Performance")
            print(f"{'='*50}")
            xirr = result.get("xirr")
            if xirr is not None:
                print(f"  XIRR: {xirr:.2%}")
            cagr = result.get("annual_return")
            if cagr is not None:
                print(f"  CAGR: {cagr:.2%}")
            print(f"  Years held: {result.get('years_held')}")
            bd = result.get("buffett_distance", {})
            if bd.get("score") is not None:
                print(f"  Buffett Distance: {bd['score']}")
                print(f"  Rating: {bd['rating']}")
                print(f"  {bd['description']}")
            print()
            export_performance_to_excel(result)
    else:
        # Demo with sample data
        print("Demo: XIRR Calculation")
        print("-" * 40)

        # Example: invest $1000, receive $1100 after 1 year → XIRR ≈ 10%
        cashflows = [-1000, 1100]
        dates = [date(2023, 1, 1), date(2024, 1, 1)]
        xirr = calc_xirr(cashflows, dates)
        print(f"  Invest $1000, receive $1100 after 1 year")
        print(f"  XIRR = {xirr:.4%}" if xirr else "  XIRR = N/A")

        # Example: more complex
        cashflows2 = [-10000, 500, 500, 500, 12000]
        dates2 = [
            date(2020, 1, 1), date(2021, 1, 1), date(2022, 1, 1),
            date(2023, 1, 1), date(2024, 1, 1),
        ]
        xirr2 = calc_xirr(cashflows2, dates2)
        print(f"\n  Invest $10000, receive $500/yr dividends + sell at $12000 after 4 years")
        print(f"  XIRR = {xirr2:.4%}" if xirr2 else "  XIRR = N/A")

        # Buffett distance baseline test: 8 years, 12% return → score ≈ 100
        print("\nDemo: Buffett Distance")
        print("-" * 40)
        bd = calc_buffett_distance(8, 0.12)
        print(f"  8 years, 12% return → Score: {bd['score']}, Rating: {bd['rating']}")

        bd2 = calc_buffett_distance(20, 0.20)
        print(f"  20 years, 20% return → Score: {bd2['score']}, Rating: {bd2['rating']}")

        bd3 = calc_buffett_distance(5, 0.08)
        print(f"  5 years, 8% return → Score: {bd3['score']}, Rating: {bd3['rating']}")


if __name__ == "__main__":
    main()
