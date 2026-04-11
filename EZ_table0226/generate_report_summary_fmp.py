"""
generate_report_summary_fmp.py
FMP-based version of the report summary generator.
Uses fmp_client.py instead of yfinance to fetch all financial data,
then outputs report_summary.xlsx matching the 盈再表 "美股" sheet layout.

Usage:
    python generate_report_summary_fmp.py              # uses default ticker
    python generate_report_summary_fmp.py AAPL          # specify ticker
    python generate_report_summary_fmp.py 9022.T JPY    # specify ticker + currency
"""

import sys
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook

import fmp_client

# ========= Configuration =========
DEFAULT_TICKER = "AAPL"
DEFAULT_CURRENCY = None  # Auto-detect from financialCurrency; set e.g. "JPY" to override

OUTPUT_FILE = "report_summary.xlsx"
SHEET_NAME = "美股"


# =====================================================================
# Helper
# =====================================================================

def _sanitize(df):
    """Clean DataFrame for Excel output."""
    df = df.copy()
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.replace([np.inf, -np.inf], np.nan)
    return df


def _safe_val(val):
    """Return a safe value for writing to Excel (no NaN/inf, no formula injection)."""
    if val is None:
        return ""
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return ""
    # Prevent Excel formula injection from data values
    if isinstance(val, str) and len(val) > 0 and val[0] in ("=", "+", "-", "@"):
        return "'" + val
    return val


# =====================================================================
# Excel Writing Functions (identical to yfinance version)
# =====================================================================

def write_front_page(ws, profile, eps_data, share_cap, forex, net_income_ttm):
    """Write front-page cells: A1, A2, I1, F15, K3, Y16, Y23, Y24, W25, E9, K10-K12."""
    # A1: {Company Name} {Ticker} ({Exchange}) : {Sector}*{Industry}
    company_line = (
        f"{profile.get('longName', '')} {profile.get('symbol', '')} "
        f"({profile.get('exchange', '')}) : "
        f"{profile.get('sector', '')}*{profile.get('industry', '')}"
    )
    ws.cell(row=1, column=1, value=company_line)

    # A2: Ticker symbol
    ws.cell(row=2, column=1, value=profile.get("symbol", ""))

    # I1: Today's date
    ws.cell(row=1, column=9, value=datetime.today().strftime("%Y-%m-%d"))

    # F15: Exchange rate
    ws.cell(row=15, column=6, value=_safe_val(forex.get("rate", 1.0)))

    # K3: Current stock price
    ws.cell(row=3, column=11, value=_safe_val(profile.get("currentPrice")))

    # Y16: Company description (longBusinessSummary)
    ws.cell(row=16, column=25, value=_safe_val(profile.get("longBusinessSummary", "")))

    # Y23: Market Cap (in millions)
    mc = profile.get("marketCap")
    if mc is not None:
        ws.cell(row=23, column=25, value=round(mc / 1_000_000, 2))
    else:
        ws.cell(row=23, column=25, value="")

    # Y24: Stock price
    ws.cell(row=24, column=25, value=_safe_val(profile.get("currentPrice")))

    # W25: Financial currency code
    ws.cell(row=25, column=23, value=_safe_val(profile.get("financialCurrency", "USD")))

    # E9: Net Income TTM (millions)
    if net_income_ttm is not None:
        ws.cell(row=9, column=5, value=round(net_income_ttm / 1_000_000, 2))
    else:
        ws.cell(row=9, column=5, value="")

    # K10: 12 (constant — months)
    ws.cell(row=10, column=11, value=12)

    # K11: Payout ratio
    ws.cell(row=11, column=11, value=_safe_val(profile.get("payoutRatio")))

    # K12: 0.25 (dividend tax default)
    ws.cell(row=12, column=11, value=0.25)


def write_financial_table(ws, df, start_col, section_title, section_num):
    """
    Write a financial table (IS/BS/CFS) to the worksheet.
    Row 1: section header
    Row 2: "Item" + date columns
    Row 3+: data rows
    """
    if df.empty:
        ws.cell(row=1, column=start_col, value=f"{section_num} / 9 {section_title}")
        return

    # Row 1: Section header
    ws.cell(row=1, column=start_col, value=f"{section_num} / 9 {section_title}")

    # Row 2: Column headers
    ws.cell(row=2, column=start_col, value="Item")
    for ci, col_name in enumerate(df.columns, start=1):
        ws.cell(row=2, column=start_col + ci, value=col_name)

    # Row 3+: Data
    for ri, (item_name, row_data) in enumerate(df.iterrows(), start=3):
        ws.cell(row=ri, column=start_col, value=item_name)
        for ci, col_name in enumerate(df.columns, start=1):
            val = row_data[col_name]
            ws.cell(row=ri, column=start_col + ci, value=_safe_val(val))


def write_unusual_items(ws, unusual_df, start_col, data_row_count):
    """Append unusual items below the main IS quarterly table."""
    if unusual_df.empty:
        return
    # Place separator + unusual items after the main data
    sep_row = 3 + data_row_count + 1
    ws.cell(row=sep_row, column=start_col, value="[Unusual / Special Items]")

    for ri, (item_name, row_data) in enumerate(unusual_df.iterrows(), start=sep_row + 1):
        ws.cell(row=ri, column=start_col, value=item_name)
        for ci, col_name in enumerate(unusual_df.columns, start=1):
            val = row_data[col_name]
            ws.cell(row=ri, column=start_col + ci, value=_safe_val(val))


def write_price_history(ws, combined_df, start_col):
    """
    Write historical price section with adjusted close and corporate actions.
    Row 1 (BO1): section header
    Row 4 (BO4:BT4+): column headers
    Row 5+: combined monthly price + corporate action data (newest first)
    """
    ws.cell(row=1, column=start_col, value="5 / 9 Historical stock price (adj)")

    if combined_df.empty:
        return

    # Row 4: Column headers
    headers = ["Date", "Open", "High", "Low", "Close", "Adj Close",
               "Volume", "Dividends", "Stock Splits"]
    for ci, h in enumerate(headers):
        ws.cell(row=4, column=start_col + ci, value=h)

    # Row 5+: Data rows (already sorted descending)
    data_cols = ["Open", "High", "Low", "Close", "Adj Close",
                 "Volume", "Dividends", "Stock Splits"]
    row_offset = 5
    for ri, (dt_idx, row_data) in enumerate(combined_df.iterrows()):
        r = row_offset + ri
        # Date column
        if isinstance(dt_idx, str):
            ws.cell(row=r, column=start_col, value=dt_idx)
        else:
            ws.cell(row=r, column=start_col, value=dt_idx.strftime("%Y/%m/%d"))
        # Data columns
        for ci, col in enumerate(data_cols, start=1):
            if col in combined_df.columns:
                ws.cell(row=r, column=start_col + ci, value=_safe_val(row_data[col]))


def write_profile_section(ws, profile, start_col):
    """Write company profile raw data section (6/9)."""
    ws.cell(row=1, column=start_col, value="6 / 9 Company Profile")

    items = [
        ("Company Name", profile.get("longName")),
        ("Ticker", profile.get("symbol")),
        ("Exchange", profile.get("exchange")),
        ("Sector", profile.get("sector")),
        ("Industry", profile.get("industry")),
        ("Country", profile.get("country")),
        ("Currency", profile.get("currency")),
        ("Financial Currency", profile.get("financialCurrency")),
        ("Current Price", profile.get("currentPrice")),
        ("Market Cap", profile.get("marketCap")),
        ("Enterprise Value", profile.get("enterpriseValue")),
        ("Trailing PE", profile.get("trailingPE")),
        ("Forward PE", profile.get("forwardPE")),
        ("EPS (TTM)", profile.get("trailingEps")),
        ("Dividend Yield", profile.get("dividendYield")),
        ("Payout Ratio", profile.get("payoutRatio")),
        ("Beta", profile.get("beta")),
        ("52 Week High", profile.get("fiftyTwoWeekHigh")),
        ("52 Week Low", profile.get("fiftyTwoWeekLow")),
        ("Revenue (TTM)", profile.get("totalRevenue")),
        ("Net Income (TTM)", profile.get("netIncomeToCommon")),
        ("Website", profile.get("website")),
        ("Description", profile.get("longBusinessSummary")),
    ]
    for ri, (item, value) in enumerate(items, start=2):
        ws.cell(row=ri, column=start_col, value=item)
        ws.cell(row=ri, column=start_col + 1, value=_safe_val(value))


def write_forex_section(ws, forex, start_col):
    """Write exchange rate section (7/9)."""
    ws.cell(row=1, column=start_col, value="7 / 9 Foreign exchange rate")

    items = [
        ("Currency Pair", forex.get("pair", "")),
        ("Exchange Rate (Close)", forex.get("rate")),
        ("Open", forex.get("open")),
        ("High", forex.get("high")),
        ("Low", forex.get("low")),
    ]
    for ri, (item, value) in enumerate(items, start=2):
        ws.cell(row=ri, column=start_col, value=item)
        ws.cell(row=ri, column=start_col + 1, value=_safe_val(value))


def write_market_cap_section(ws, profile, share_cap, eps_data, start_col):
    """Write market cap / valuation section (9/9)."""
    ws.cell(row=1, column=start_col, value="9 / 9 Market capitalization")

    mc = profile.get("marketCap")
    ev = profile.get("enterpriseValue")
    items = [
        ("Market Cap ($M)", round(mc / 1_000_000, 2) if mc else ""),
        ("Enterprise Value ($M)", round(ev / 1_000_000, 2) if ev else ""),
        ("Trailing PE", profile.get("trailingPE")),
        ("Forward PE", profile.get("forwardPE")),
        ("EPS (TTM)", profile.get("trailingEps")),
        ("Dividend Yield", profile.get("dividendYield")),
        ("Payout Ratio", profile.get("payoutRatio")),
        ("Beta", profile.get("beta")),
        ("52 Week High", profile.get("fiftyTwoWeekHigh")),
        ("52 Week Low", profile.get("fiftyTwoWeekLow")),
        ("Revenue (TTM)", profile.get("totalRevenue")),
        ("Net Income (TTM)", profile.get("netIncomeToCommon")),
        ("Shares Outstanding", share_cap.get("Shares Outstanding")),
        ("Float Shares", share_cap.get("Float Shares")),
        ("Implied Shares Outstanding", share_cap.get("Implied Shares Outstanding")),
        ("Held by Insiders (%)", share_cap.get("Held by Insiders (%)")),
        ("Held by Institutions (%)", share_cap.get("Held by Institutions (%)")),
        ("Short Shares", share_cap.get("Short Shares")),
        ("Short Prior Month", share_cap.get("Short Prior Month")),
        ("Short % of Shares Outstanding", share_cap.get("Short % of Shares Outstanding")),
    ]

    # Append EPS earnings data
    if eps_data.get("ttm_eps") is not None:
        items.append(("TTM EPS", eps_data["ttm_eps"]))
    if eps_data.get("annual"):
        items.append(("[Annual Earnings]", ""))
        for year, val in eps_data["annual"].items():
            items.append((f"Earnings {year}", val))
    if eps_data.get("quarterly"):
        items.append(("[Quarterly Earnings]", ""))
        for date_str, val in eps_data["quarterly"].items():
            items.append((f"Earnings {date_str}", val))

    for ri, (item, value) in enumerate(items, start=2):
        ws.cell(row=ri, column=start_col, value=item)
        ws.cell(row=ri, column=start_col + 1, value=_safe_val(value))


# =====================================================================
# Main
# =====================================================================

def main():
    # Parse CLI args
    ticker_symbol = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_TICKER
    currency_override = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_CURRENCY

    print(f"{'='*60}")
    print(f"  Report Summary Generator (FMP)")
    print(f"  Ticker: {ticker_symbol}")
    print(f"{'='*60}")

    # ---- Fetch all data via FMP ----
    print("1/10 Fetching quarterly income statement...")
    is_q_result = fmp_client.fetch_quarterly_income(ticker_symbol)
    if isinstance(is_q_result, tuple):
        is_q, unusual_q = is_q_result
    else:
        is_q, unusual_q = is_q_result, pd.DataFrame()

    print("2/10 Fetching quarterly balance sheet...")
    bs_q = fmp_client.fetch_quarterly_balance_sheet(ticker_symbol)

    print("3/10 Fetching annual income statement...")
    is_y = fmp_client.fetch_annual_income(ticker_symbol)

    print("4/10 Fetching annual balance sheet...")
    bs_y = fmp_client.fetch_annual_balance_sheet(ticker_symbol)

    print("5/10 Fetching annual cash flow statement...")
    cfs_y = fmp_client.fetch_annual_cashflow(ticker_symbol)

    print("6/10 Fetching company profile...")
    profile = fmp_client.fetch_company_profile(ticker_symbol)

    print("7/10 Fetching share capital data...")
    share_cap = fmp_client.fetch_share_capital(ticker_symbol)

    print("8/10 Fetching EPS / earnings data...")
    eps_data = fmp_client.fetch_eps_earnings(ticker_symbol)

    print("9/10 Fetching historical prices (adjusted) & actions...")
    combined_prices = fmp_client.fetch_historical_prices_adj(ticker_symbol)

    # Determine currency and fetch exchange rate
    fin_currency = currency_override or profile.get("financialCurrency", "USD")
    print(f"10/10 Fetching exchange rate (USD/{fin_currency})...")
    forex = fmp_client.fetch_exchange_rate(fin_currency)

    # ---- Compute Net Income TTM for front page ----
    net_income_ttm = None
    if not is_q.empty:
        for label in ["netIncome", "netIncomeDeducted"]:
            if label in is_q.index and "TTM" in is_q.columns:
                val = is_q.loc[label, "TTM"]
                if pd.notna(val):
                    net_income_ttm = val
                    break
    # Fallback to profile netIncomeToCommon
    if net_income_ttm is None and profile.get("netIncomeToCommon") is not None:
        net_income_ttm = profile["netIncomeToCommon"]

    # ---- Column positions (美股 sheet layout) ----
    COL_IS_Q = 31   # AE — 1/9 Quarterly IS
    COL_BS_Q = 40   # AN — 2/9 Quarterly BS
    COL_IS_Y = 49   # AW — 3/9 Annual IS
    COL_BS_Y = 58   # BF — 4/9 Annual BS
    COL_PRICE = 67  # BO — 5/9 Historical Price
    COL_PROF = 76   # BX — 6/9 Company Profile
    COL_FX = 79     # CA — 7/9 Exchange Rate
    COL_CFS = 83    # CE — 8/9 Annual CFS
    COL_MCAP = 92   # CN — 9/9 Market Cap

    # ---- Create workbook ----
    print("\nWriting Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Front page cells
    write_front_page(ws, profile, eps_data, share_cap, forex, net_income_ttm)

    # 1/9 Quarterly Income Statement
    write_financial_table(ws, is_q, COL_IS_Q, "Quarterly Income Statement", 1)
    if not unusual_q.empty:
        write_unusual_items(ws, unusual_q, COL_IS_Q, len(is_q))

    # 2/9 Quarterly Balance Sheet
    write_financial_table(ws, bs_q, COL_BS_Q, "Quarterly Balance Sheet", 2)

    # 3/9 Annual Income Statement
    write_financial_table(ws, is_y, COL_IS_Y, "Annual Income Statement", 3)

    # 4/9 Annual Balance Sheet
    write_financial_table(ws, bs_y, COL_BS_Y, "Annual Balance Sheet", 4)

    # 5/9 Historical Stock Price (adj)
    write_price_history(ws, combined_prices, COL_PRICE)

    # 6/9 Company Profile
    write_profile_section(ws, profile, COL_PROF)

    # 7/9 Exchange Rate
    write_forex_section(ws, forex, COL_FX)

    # 8/9 Annual Cash Flow
    write_financial_table(ws, cfs_y, COL_CFS, "Annual Cash Flow", 8)

    # 9/9 Market Capitalization
    write_market_cap_section(ws, profile, share_cap, eps_data, COL_MCAP)

    # ---- Save ----
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved: {OUTPUT_FILE}")
    print(f"   Sheet: {SHEET_NAME}")
    print(f"   Sections written:")
    print(f"     1/9 Quarterly IS  → col AE ({COL_IS_Q})")
    print(f"     2/9 Quarterly BS  → col AN ({COL_BS_Q})")
    print(f"     3/9 Annual IS     → col AW ({COL_IS_Y})")
    print(f"     4/9 Annual BS     → col BF ({COL_BS_Y})")
    print(f"     5/9 Price History (adj) → col BO ({COL_PRICE}), headers BO4:BW4")
    print(f"     6/9 Profile       → col BX ({COL_PROF})")
    print(f"     7/9 Exchange Rate → col CA ({COL_FX})")
    print(f"     8/9 Annual CFS    → col CE ({COL_CFS})")
    print(f"     9/9 Market Cap    → col CN ({COL_MCAP})")
    print(f"\n   Front-page cells: A1, A2, I1, F15, K3, Y16, Y23, Y24, W25, E9, K10-K12")


if __name__ == "__main__":
    main()
