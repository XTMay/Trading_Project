"""
financial_calculator.py
Core financial metrics engine for 盈再表 (Profit Reinvestment Rate Analysis).

Implements the front-end calculations (VBA Module1-5 Step 10-11):
  - ROE% (Return on Equity)
  - 盈再率% (Profit Reinvestment Rate)
  - 常利 (Recurring Profit)
  - 常 EPS (Recurring EPS)
  - 配息% (Payout Ratio)
  - 預期報酬 (Expected Return)
  - Price Zones (便宜价/合理价/昂贵价)
  - 還原股價 (Restored / Dividend-Adjusted Price)

Usage:
    python financial_calculator.py              # uses default ticker AAPL
    python financial_calculator.py AAPL         # specify ticker
    python financial_calculator.py 2330.TW TWD  # specify ticker + currency
"""

import sys
import math
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook

from generate_report_summary import (
    _sanitize,
    _safe_val,
    fetch_annual_income,
    fetch_annual_balance_sheet,
    fetch_annual_cashflow,
    fetch_company_profile,
    fetch_share_capital,
    fetch_historical_prices_adj,
    fetch_exchange_rate,
)

# =====================================================================
# Core Metric Calculations
# =====================================================================


def calc_roe(is_y, bs_y):
    """Calculate ROE% = Net Income / Stockholders Equity per year.

    Args:
        is_y: Annual income statement DataFrame (rows=items, cols=years+TTM).
        bs_y: Annual balance sheet DataFrame.

    Returns:
        dict: {year_label: roe_value} where roe_value is a float (e.g. 0.25 = 25%).
    """
    ni_row = None
    for label in ["Net Income", "Net Income Common Stockholders"]:
        if label in is_y.index:
            ni_row = label
            break

    eq_row = None
    for label in ["Stockholders Equity", "Total Equity Gross Minority Interest",
                   "Stockholders' Equity"]:
        if label in bs_y.index:
            eq_row = label
            break

    if ni_row is None or eq_row is None:
        return {}

    common_cols = sorted(set(is_y.columns) & set(bs_y.columns))
    result = {}
    for col in common_cols:
        ni = pd.to_numeric(is_y.loc[ni_row, col], errors="coerce")
        eq = pd.to_numeric(bs_y.loc[eq_row, col], errors="coerce")
        if pd.notna(ni) and pd.notna(eq) and eq != 0:
            result[col] = ni / eq
    return result


def calc_reinvestment_rate(cfs_y, is_y):
    """Calculate 盈再率% = |Capital Expenditure| / Net Income per year.

    Capital Expenditure is usually negative in yfinance; we take abs().

    Args:
        cfs_y: Annual cash flow statement DataFrame.
        is_y: Annual income statement DataFrame.

    Returns:
        dict: {year_label: reinvestment_rate} (e.g. 0.30 = 30%).
    """
    capex_row = None
    for label in ["Capital Expenditure", "CapitalExpenditure"]:
        if label in cfs_y.index:
            capex_row = label
            break

    ni_row = None
    for label in ["Net Income", "Net Income Common Stockholders"]:
        if label in is_y.index:
            ni_row = label
            break

    if capex_row is None or ni_row is None:
        return {}

    common_cols = sorted(set(cfs_y.columns) & set(is_y.columns))
    result = {}
    for col in common_cols:
        capex = pd.to_numeric(cfs_y.loc[capex_row, col], errors="coerce")
        ni = pd.to_numeric(is_y.loc[ni_row, col], errors="coerce")
        if pd.notna(capex) and pd.notna(ni) and ni > 0:
            result[col] = abs(capex) / ni
    return result


def calc_recurring_profit(is_y):
    """Calculate 常利 = Net Income - Unusual Items per year.

    Unusual items are identified by keywords in the row index.

    Args:
        is_y: Annual income statement DataFrame.

    Returns:
        dict: {year_label: recurring_profit_value}.
    """
    ni_row = None
    for label in ["Net Income", "Net Income Common Stockholders"]:
        if label in is_y.index:
            ni_row = label
            break
    if ni_row is None:
        return {}

    keywords = ["unusual", "special", "restructuring", "non recurring",
                 "nonrecurring", "impairment", "write off", "writeoff"]
    unusual_mask = is_y.index.str.lower().str.contains("|".join(keywords))
    unusual_rows = is_y.index[unusual_mask].tolist()

    result = {}
    for col in is_y.columns:
        ni = pd.to_numeric(is_y.loc[ni_row, col], errors="coerce")
        if pd.isna(ni):
            continue
        unusual_total = 0.0
        for row in unusual_rows:
            val = pd.to_numeric(is_y.loc[row, col], errors="coerce")
            if pd.notna(val):
                unusual_total += val
        result[col] = ni - unusual_total
    return result


def calc_recurring_eps(recurring_profit, shares_outstanding):
    """Calculate 常 EPS = Recurring Profit / Shares Outstanding.

    Args:
        recurring_profit: dict {year: profit_value}.
        shares_outstanding: int or float, total shares outstanding.

    Returns:
        dict: {year_label: recurring_eps_value}.
    """
    if not shares_outstanding or shares_outstanding <= 0:
        return {}
    return {
        year: profit / shares_outstanding
        for year, profit in recurring_profit.items()
        if profit is not None
    }


def calc_payout_ratio(ticker, is_y):
    """Calculate 配息% = Annual Dividends / Net Income.

    Fetches dividend data from ticker.actions.

    Args:
        ticker: yfinance Ticker object.
        is_y: Annual income statement DataFrame.

    Returns:
        dict: {year_label: payout_ratio}.
    """
    try:
        actions = ticker.actions
        if actions is None or actions.empty:
            return {}
    except Exception:
        return {}

    dividends = actions[actions["Dividends"] > 0]["Dividends"]
    if dividends.empty:
        return {}

    # Aggregate dividends by year
    div_by_year = dividends.groupby(dividends.index.year).sum()

    ni_row = None
    for label in ["Net Income", "Net Income Common Stockholders"]:
        if label in is_y.index:
            ni_row = label
            break
    if ni_row is None:
        return {}

    # Get shares outstanding for per-share conversion
    try:
        shares = ticker.info.get("sharesOutstanding")
    except Exception:
        shares = None

    result = {}
    for col in is_y.columns:
        if col == "TTM":
            continue
        try:
            year = int(col)
        except (ValueError, TypeError):
            continue
        ni = pd.to_numeric(is_y.loc[ni_row, col], errors="coerce")
        if pd.isna(ni) or ni <= 0:
            continue
        if year in div_by_year.index:
            total_div = div_by_year.loc[year]
            if shares and shares > 0:
                # total_div is per-share; convert to total
                total_div_amount = total_div * shares
            else:
                total_div_amount = total_div
            ratio = total_div_amount / ni
            if 0 <= ratio <= 2.0:  # sanity check
                result[col] = ratio
    return result


def calc_expected_return(roe, reinv_rate):
    """Calculate 預期報酬 = ROE × (1 - 盈再率).

    Uses the average of available years for both metrics.

    Args:
        roe: dict {year: roe_value}.
        reinv_rate: dict {year: reinvestment_rate_value}.

    Returns:
        float: expected annual return rate, or None.
    """
    common_years = sorted(set(roe.keys()) & set(reinv_rate.keys()))
    if not common_years:
        # Fall back to average ROE if no reinvestment data
        if roe:
            vals = [v for v in roe.values() if v is not None and 0 < v < 1]
            return sum(vals) / len(vals) if vals else None
        return None

    returns = []
    for year in common_years:
        r = roe[year]
        ri = reinv_rate[year]
        if r is not None and ri is not None and 0 < r < 1 and 0 <= ri < 1:
            returns.append(r * (1 - ri))

    return sum(returns) / len(returns) if returns else None


def calc_price_zones(recurring_eps_latest, expected_return):
    """Calculate price zones: 便宜价 / 合理价 / 昂贵价.

    Based on the 盈再表 methodology:
      - 便宜价 (cheap)  = recurring_eps / expected_return × discount_factor
      - 合理价 (fair)    = recurring_eps / expected_return
      - 昂贵价 (expensive) = recurring_eps / expected_return × premium_factor

    Args:
        recurring_eps_latest: Most recent recurring EPS value.
        expected_return: Expected annual return rate (e.g. 0.15).

    Returns:
        dict with keys: cheap_price, fair_price, expensive_price.
    """
    if (recurring_eps_latest is None or expected_return is None
            or expected_return <= 0 or recurring_eps_latest <= 0):
        return {"cheap_price": None, "fair_price": None, "expensive_price": None}

    fair = recurring_eps_latest / expected_return
    return {
        "cheap_price": round(fair * 0.6, 2),
        "fair_price": round(fair, 2),
        "expensive_price": round(fair * 1.4, 2),
    }


def calc_restored_price(prices_df):
    """Calculate 還原股價: cumulative reverse-adjust for dividends and splits.

    Processes historical price data to compute prices as if all dividends
    were reinvested and no splits occurred.

    Args:
        prices_df: DataFrame from fetch_historical_prices_adj (sorted newest first).

    Returns:
        pd.Series: restored prices indexed by date (newest first).
    """
    if prices_df is None or prices_df.empty:
        return pd.Series(dtype=float)

    # Work oldest-first for cumulative calculation
    df = prices_df.sort_index(ascending=True).copy()

    if "Close" not in df.columns:
        return pd.Series(dtype=float)

    close = df["Close"].apply(pd.to_numeric, errors="coerce")
    dividends = df.get("Dividends", pd.Series(0, index=df.index))
    dividends = dividends.apply(pd.to_numeric, errors="coerce").fillna(0)
    splits = df.get("Stock Splits", pd.Series(0, index=df.index))
    splits = splits.apply(pd.to_numeric, errors="coerce").fillna(0)

    restored = []
    cum_div = 0.0
    cum_split_factor = 1.0

    for i in range(len(df)):
        c = close.iloc[i]
        d = dividends.iloc[i]
        s = splits.iloc[i]

        if pd.notna(d) and d > 0:
            cum_div += d

        if pd.notna(s) and s > 0:
            cum_split_factor *= s
            cum_div *= s  # adjust accumulated dividends for split

        if pd.notna(c):
            restored_val = (c + cum_div) / cum_split_factor
            restored.append(restored_val)
        else:
            restored.append(np.nan)

    result = pd.Series(restored, index=df.index)
    return result.sort_index(ascending=False)


def generate_analysis(ticker_symbol, currency=None):
    """Master function: fetch data + compute all metrics.

    Args:
        ticker_symbol: Stock ticker (e.g. "AAPL", "2330.TW").
        currency: Optional currency override (e.g. "TWD", "JPY").

    Returns:
        dict with all computed metrics and raw data.
    """
    print(f"{'='*60}")
    print(f"  Financial Analysis: {ticker_symbol}")
    print(f"{'='*60}")

    ticker = yf.Ticker(ticker_symbol)

    # Fetch data
    print("  Fetching annual income statement...")
    is_y = fetch_annual_income(ticker)
    print("  Fetching annual balance sheet...")
    bs_y = fetch_annual_balance_sheet(ticker)
    print("  Fetching annual cash flow...")
    cfs_y = fetch_annual_cashflow(ticker)
    print("  Fetching company profile...")
    profile = fetch_company_profile(ticker)
    print("  Fetching share capital...")
    share_cap = fetch_share_capital(ticker)
    print("  Fetching historical prices...")
    prices = fetch_historical_prices_adj(ticker)

    fin_currency = currency or profile.get("financialCurrency", "USD")
    forex = fetch_exchange_rate(fin_currency)

    shares = share_cap.get("Shares Outstanding") or 0

    # Compute metrics
    print("  Computing ROE%...")
    roe = calc_roe(is_y, bs_y)
    print("  Computing reinvestment rate...")
    reinv = calc_reinvestment_rate(cfs_y, is_y)
    print("  Computing recurring profit...")
    rec_profit = calc_recurring_profit(is_y)
    print("  Computing recurring EPS...")
    rec_eps = calc_recurring_eps(rec_profit, shares)
    print("  Computing payout ratio...")
    payout = calc_payout_ratio(ticker, is_y)
    print("  Computing expected return...")
    exp_return = calc_expected_return(roe, reinv)

    # Get latest recurring EPS for price zones
    rec_eps_latest = None
    if rec_eps:
        sorted_years = sorted(rec_eps.keys())
        # Prefer non-TTM latest year
        non_ttm = [y for y in sorted_years if y != "TTM"]
        key = non_ttm[-1] if non_ttm else sorted_years[-1]
        rec_eps_latest = rec_eps[key]

    print("  Computing price zones...")
    zones = calc_price_zones(rec_eps_latest, exp_return)

    print("  Computing restored prices...")
    restored = calc_restored_price(prices)

    current_price = profile.get("currentPrice")

    # Determine price zone status
    zone_status = "N/A"
    if current_price and zones["cheap_price"]:
        if current_price <= zones["cheap_price"]:
            zone_status = "便宜 (Cheap)"
        elif current_price <= zones["fair_price"]:
            zone_status = "合理 (Fair)"
        elif current_price <= zones["expensive_price"]:
            zone_status = "偏贵 (Pricey)"
        else:
            zone_status = "昂贵 (Expensive)"

    result = {
        "ticker": ticker_symbol,
        "company": profile.get("longName", ""),
        "current_price": current_price,
        "currency": profile.get("currency", ""),
        "financial_currency": fin_currency,
        "exchange_rate": forex.get("rate", 1.0),
        "shares_outstanding": shares,
        "roe": roe,
        "reinvestment_rate": reinv,
        "recurring_profit": rec_profit,
        "recurring_eps": rec_eps,
        "payout_ratio": payout,
        "expected_return": exp_return,
        "price_zones": zones,
        "zone_status": zone_status,
        "restored_prices": restored,
        "profile": profile,
    }

    _print_summary(result)
    return result


def _print_summary(result):
    """Print a formatted summary of analysis results."""
    print(f"\n{'='*60}")
    print(f"  {result['company']} ({result['ticker']})")
    print(f"  Current Price: {result['current_price']} {result['currency']}")
    print(f"{'='*60}")

    print("\n  ROE% by year:")
    for year, val in sorted(result["roe"].items()):
        print(f"    {year}: {val:.2%}")

    print("\n  Reinvestment Rate% by year:")
    for year, val in sorted(result["reinvestment_rate"].items()):
        print(f"    {year}: {val:.2%}")

    print("\n  Recurring Profit ($M) by year:")
    for year, val in sorted(result["recurring_profit"].items()):
        print(f"    {year}: {val/1e6:,.2f}")

    print("\n  Recurring EPS by year:")
    for year, val in sorted(result["recurring_eps"].items()):
        print(f"    {year}: {val:.4f}")

    if result["expected_return"] is not None:
        print(f"\n  Expected Return: {result['expected_return']:.2%}")

    zones = result["price_zones"]
    if zones["fair_price"] is not None:
        print(f"\n  Price Zones:")
        print(f"    Cheap  (便宜价): {zones['cheap_price']}")
        print(f"    Fair   (合理价): {zones['fair_price']}")
        print(f"    Expensive (昂贵价): {zones['expensive_price']}")
        print(f"    Status: {result['zone_status']}")

    print()


def export_analysis_to_excel(result, output_path="financial_analysis.xlsx"):
    """Write analysis results to an Excel file.

    Args:
        result: dict from generate_analysis().
        output_path: Output file path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

    # Header
    ws.cell(row=1, column=1, value=f"{result['company']} ({result['ticker']})")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.today().strftime('%Y-%m-%d')}")
    ws.cell(row=2, column=3, value=f"Price: {_safe_val(result['current_price'])}")

    # Summary metrics
    row = 4
    ws.cell(row=row, column=1, value="Expected Return")
    ws.cell(row=row, column=2,
            value=_safe_val(result["expected_return"]))

    row += 1
    ws.cell(row=row, column=1, value="Cheap Price (便宜价)")
    ws.cell(row=row, column=2,
            value=_safe_val(result["price_zones"]["cheap_price"]))
    row += 1
    ws.cell(row=row, column=1, value="Fair Price (合理价)")
    ws.cell(row=row, column=2,
            value=_safe_val(result["price_zones"]["fair_price"]))
    row += 1
    ws.cell(row=row, column=1, value="Expensive Price (昂贵价)")
    ws.cell(row=row, column=2,
            value=_safe_val(result["price_zones"]["expensive_price"]))
    row += 1
    ws.cell(row=row, column=1, value="Zone Status")
    ws.cell(row=row, column=2, value=result["zone_status"])

    # Yearly metrics table
    row += 2
    ws.cell(row=row, column=1, value="Year")
    ws.cell(row=row, column=2, value="ROE%")
    ws.cell(row=row, column=3, value="Reinv Rate%")
    ws.cell(row=row, column=4, value="Recurring Profit ($M)")
    ws.cell(row=row, column=5, value="Recurring EPS")
    ws.cell(row=row, column=6, value="Payout Ratio")

    all_years = sorted(set(
        list(result["roe"].keys()) +
        list(result["reinvestment_rate"].keys()) +
        list(result["recurring_profit"].keys())
    ))

    for year in all_years:
        row += 1
        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2,
                value=_safe_val(result["roe"].get(year)))
        ws.cell(row=row, column=3,
                value=_safe_val(result["reinvestment_rate"].get(year)))
        rp = result["recurring_profit"].get(year)
        ws.cell(row=row, column=4,
                value=_safe_val(round(rp / 1e6, 2) if rp else None))
        ws.cell(row=row, column=5,
                value=_safe_val(result["recurring_eps"].get(year)))
        ws.cell(row=row, column=6,
                value=_safe_val(result["payout_ratio"].get(year)))

    wb.save(output_path)
    print(f"Saved: {output_path}")


# =====================================================================
# CLI Entry Point
# =====================================================================

def main():
    ticker_symbol = sys.argv[1] if len(sys.argv) > 1 else "AAPL"
    currency = sys.argv[2] if len(sys.argv) > 2 else None

    result = generate_analysis(ticker_symbol, currency)

    output_file = f"financial_analysis_{ticker_symbol.replace('.', '_')}.xlsx"
    export_analysis_to_excel(result, output_file)


if __name__ == "__main__":
    main()
