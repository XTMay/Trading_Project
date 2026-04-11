"""
portfolio_manager.py
Portfolio Management Engine for 盈再表.

Implements VBA Module7 logic:
  - Load holdings from Excel
  - Update prices via yfinance
  - Calculate expected returns for each holding
  - Rank holdings by expected return
  - Check duplicate holdings across markets
  - Generate portfolio summary
  - Export portfolio report to Excel

Usage:
    python portfolio_manager.py                     # demo
    python portfolio_manager.py portfolio.xlsx      # load from Excel
"""

import sys
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook

from generate_report_summary import _safe_val

# =====================================================================
# Holdings Management
# =====================================================================


def load_holdings(excel_path, sheet_name=None):
    """Read holdings from Excel file.

    Expected columns (flexible matching):
      - code/ticker/股票代号: Stock ticker symbol
      - shares/股数/持股: Number of shares
      - buy_price/成本价/买入价: Average buy price
      - market/市场: Market identifier (US/TW/HK/CN)

    Args:
        excel_path: Path to Excel file.
        sheet_name: Optional sheet name.

    Returns:
        pd.DataFrame with standardized columns:
            code, shares, buy_price, market, current_price (initially NaN).
    """
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"Error reading {excel_path}: {e}")
        return pd.DataFrame()

    # Flexible column mapping
    col_map = {}
    for col in df.columns:
        c = str(col).lower().strip()
        if c in ("code", "ticker", "symbol", "股票代号", "代号", "stock"):
            col_map[col] = "code"
        elif c in ("shares", "股数", "持股", "quantity", "qty", "数量"):
            col_map[col] = "shares"
        elif c in ("buy_price", "cost", "成本价", "买入价", "avg_price", "成本"):
            col_map[col] = "buy_price"
        elif c in ("market", "市场", "exchange", "mkt"):
            col_map[col] = "market"

    df = df.rename(columns=col_map)

    required = ["code", "shares", "buy_price"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"Missing required columns: {missing}")
        print(f"Available columns: {list(df.columns)}")
        return pd.DataFrame()

    df["code"] = df["code"].astype(str).str.strip().str.upper()
    df["shares"] = pd.to_numeric(df["shares"], errors="coerce").fillna(0)
    df["buy_price"] = pd.to_numeric(df["buy_price"], errors="coerce").fillna(0)

    if "market" not in df.columns:
        df["market"] = df["code"].apply(_detect_market)

    df["current_price"] = np.nan
    df["market_value"] = np.nan
    df["cost_value"] = df["shares"] * df["buy_price"]
    df["gain_loss"] = np.nan
    df["gain_loss_pct"] = np.nan
    df["expected_return"] = np.nan

    return df


def _detect_market(code):
    """Auto-detect market from ticker suffix."""
    code = str(code).upper()
    if code.endswith(".TW") or code.endswith(".TWO"):
        return "TW"
    elif code.endswith(".HK"):
        return "HK"
    elif code.endswith(".SS") or code.endswith(".SZ"):
        return "CN"
    elif code.endswith(".T") or code.endswith(".L") or code.endswith(".PA"):
        return "Global"
    else:
        return "US"


def update_prices(holdings):
    """Fetch latest prices via yfinance for all holdings.

    Args:
        holdings: DataFrame with 'code' column.

    Returns:
        DataFrame with updated current_price, market_value, gain_loss columns.
    """
    if holdings.empty:
        return holdings

    df = holdings.copy()
    tickers = df["code"].unique().tolist()

    print(f"  Fetching prices for {len(tickers)} stocks...")
    for ticker_code in tickers:
        try:
            t = yf.Ticker(ticker_code)
            info = t.info
            price = info.get("currentPrice") or info.get("regularMarketPrice")
            if price is None:
                hist = t.history(period="5d")
                if hist is not None and not hist.empty:
                    price = hist["Close"].iloc[-1]
        except Exception:
            price = None

        mask = df["code"] == ticker_code
        if price is not None:
            df.loc[mask, "current_price"] = price
            df.loc[mask, "market_value"] = df.loc[mask, "shares"] * price
            cost = df.loc[mask, "cost_value"]
            df.loc[mask, "gain_loss"] = df.loc[mask, "market_value"] - cost
            df.loc[mask, "gain_loss_pct"] = np.where(
                cost > 0,
                (df.loc[mask, "market_value"] - cost) / cost,
                np.nan,
            )

    return df


def calc_expected_returns(holdings):
    """Compute expected return for each holding using financial_calculator.

    Args:
        holdings: DataFrame with 'code' column.

    Returns:
        DataFrame with updated 'expected_return' column.
    """
    if holdings.empty:
        return holdings

    df = holdings.copy()

    # Import here to avoid circular dependency at module load time
    from financial_calculator import generate_analysis

    tickers_done = set()
    for idx, row in df.iterrows():
        code = row["code"]
        if code in tickers_done:
            continue

        try:
            result = generate_analysis(code)
            exp_ret = result.get("expected_return")
            if exp_ret is not None:
                df.loc[df["code"] == code, "expected_return"] = exp_ret
        except Exception as e:
            print(f"  Warning: Could not compute expected return for {code}: {e}")

        tickers_done.add(code)

    return df


def rank_holdings(holdings):
    """Sort holdings by expected return descending.

    Args:
        holdings: DataFrame with 'expected_return' column.

    Returns:
        DataFrame sorted by expected_return (highest first).
    """
    df = holdings.copy()
    df["_sort_key"] = df["expected_return"].fillna(-999)
    df = df.sort_values("_sort_key", ascending=False).drop(columns=["_sort_key"])
    df = df.reset_index(drop=True)
    return df


def check_duplicates(holdings):
    """Find same company held across multiple markets.

    E.g., BABA (US) and 9988.HK (HK) are the same company.

    Args:
        holdings: DataFrame with 'code' and 'market' columns.

    Returns:
        list of tuples: [(code1, market1, code2, market2), ...].
    """
    # Known cross-listing pairs
    known_pairs = {
        "BABA": "9988.HK",
        "JD": "9618.HK",
        "PDD": "PDD",
        "BIDU": "9888.HK",
        "NIO": "9866.HK",
        "LI": "2015.HK",
        "XPEV": "9868.HK",
        "TSM": "2330.TW",
    }

    duplicates = []
    codes = holdings["code"].tolist()

    # Check by exact code match (same code, different market)
    seen = {}
    for idx, row in holdings.iterrows():
        code = row["code"]
        market = row.get("market", "")
        base = code.split(".")[0].upper()
        key = base
        if key in seen:
            prev_idx, prev_market = seen[key]
            if prev_market != market:
                duplicates.append((
                    holdings.iloc[prev_idx]["code"],
                    prev_market,
                    code,
                    market,
                ))
        else:
            seen[key] = (idx, market)

    # Check known cross-listing pairs
    code_set = set(c.upper() for c in codes)
    for us_code, hk_code in known_pairs.items():
        if us_code.upper() in code_set and hk_code.upper() in code_set:
            duplicates.append((us_code, "US", hk_code, "HK"))

    return duplicates


def generate_portfolio_summary(holdings):
    """Generate portfolio summary statistics.

    Args:
        holdings: DataFrame with price and value columns populated.

    Returns:
        dict with summary statistics.
    """
    df = holdings.copy()

    total_cost = df["cost_value"].sum()
    total_market = df["market_value"].sum()
    total_gain = total_market - total_cost if pd.notna(total_market) else None
    total_return = (total_gain / total_cost) if total_cost > 0 and total_gain is not None else None

    # Per-market breakdown
    market_stats = {}
    for market, group in df.groupby("market"):
        market_stats[market] = {
            "count": len(group),
            "cost": group["cost_value"].sum(),
            "market_value": group["market_value"].sum(),
            "gain_loss": group["gain_loss"].sum(),
        }

    # Top/bottom performers
    valid = df[df["gain_loss_pct"].notna()].copy()
    top = valid.nlargest(5, "gain_loss_pct")[["code", "gain_loss_pct"]].to_dict("records") if len(valid) > 0 else []
    bottom = valid.nsmallest(5, "gain_loss_pct")[["code", "gain_loss_pct"]].to_dict("records") if len(valid) > 0 else []

    return {
        "total_holdings": len(df),
        "total_cost": total_cost,
        "total_market_value": total_market,
        "total_gain_loss": total_gain,
        "total_return_pct": total_return,
        "market_breakdown": market_stats,
        "top_performers": top,
        "bottom_performers": bottom,
    }


def export_portfolio(holdings, summary, output_path="portfolio_report.xlsx"):
    """Write portfolio report to Excel.

    Args:
        holdings: DataFrame with all columns populated.
        summary: dict from generate_portfolio_summary().
        output_path: Output file path.
    """
    wb = Workbook()

    # Sheet 1: Holdings Detail
    ws = wb.active
    ws.title = "Holdings"

    ws.cell(row=1, column=1, value="Portfolio Holdings Report")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.today().strftime('%Y-%m-%d')}")

    headers = [
        "Code", "Market", "Shares", "Buy Price", "Current Price",
        "Cost Value", "Market Value", "Gain/Loss", "Gain/Loss %",
        "Expected Return",
    ]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=4, column=ci, value=h)

    for ri, (_, row) in enumerate(holdings.iterrows(), start=5):
        ws.cell(row=ri, column=1, value=row.get("code", ""))
        ws.cell(row=ri, column=2, value=row.get("market", ""))
        ws.cell(row=ri, column=3, value=_safe_val(row.get("shares")))
        ws.cell(row=ri, column=4, value=_safe_val(row.get("buy_price")))
        ws.cell(row=ri, column=5, value=_safe_val(row.get("current_price")))
        ws.cell(row=ri, column=6, value=_safe_val(row.get("cost_value")))
        ws.cell(row=ri, column=7, value=_safe_val(row.get("market_value")))
        ws.cell(row=ri, column=8, value=_safe_val(row.get("gain_loss")))
        glp = row.get("gain_loss_pct")
        ws.cell(row=ri, column=9, value=_safe_val(glp))
        ws.cell(row=ri, column=10, value=_safe_val(row.get("expected_return")))

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Portfolio Summary")

    row = 3
    items = [
        ("Total Holdings", summary.get("total_holdings")),
        ("Total Cost", summary.get("total_cost")),
        ("Total Market Value", summary.get("total_market_value")),
        ("Total Gain/Loss", summary.get("total_gain_loss")),
        ("Total Return %", summary.get("total_return_pct")),
    ]
    for label, val in items:
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=_safe_val(val))
        row += 1

    # Market breakdown
    row += 1
    ws2.cell(row=row, column=1, value="Market Breakdown")
    row += 1
    ws2.cell(row=row, column=1, value="Market")
    ws2.cell(row=row, column=2, value="Count")
    ws2.cell(row=row, column=3, value="Cost")
    ws2.cell(row=row, column=4, value="Market Value")
    ws2.cell(row=row, column=5, value="Gain/Loss")
    for market, stats in summary.get("market_breakdown", {}).items():
        row += 1
        ws2.cell(row=row, column=1, value=market)
        ws2.cell(row=row, column=2, value=stats.get("count"))
        ws2.cell(row=row, column=3, value=_safe_val(stats.get("cost")))
        ws2.cell(row=row, column=4, value=_safe_val(stats.get("market_value")))
        ws2.cell(row=row, column=5, value=_safe_val(stats.get("gain_loss")))

    wb.save(output_path)
    print(f"Saved: {output_path}")


# =====================================================================
# CLI Entry Point
# =====================================================================

def main():
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
        sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
        print(f"Loading holdings from: {excel_path}")
        holdings = load_holdings(excel_path, sheet_name)

        if holdings.empty:
            print("No holdings found.")
            return

        print(f"Loaded {len(holdings)} holdings.")

        # Check for duplicates
        dups = check_duplicates(holdings)
        if dups:
            print("\nDuplicate holdings detected:")
            for d in dups:
                print(f"  {d[0]} ({d[1]}) <-> {d[2]} ({d[3]})")

        # Update prices
        print("\nUpdating prices...")
        holdings = update_prices(holdings)

        # Rank
        holdings = rank_holdings(holdings)

        # Summary
        summary = generate_portfolio_summary(holdings)

        print(f"\n{'='*50}")
        print("  Portfolio Summary")
        print(f"{'='*50}")
        print(f"  Holdings: {summary['total_holdings']}")
        tc = summary.get("total_cost")
        if tc:
            print(f"  Total Cost: {tc:,.2f}")
        tmv = summary.get("total_market_value")
        if tmv and not np.isnan(tmv):
            print(f"  Total Market Value: {tmv:,.2f}")
        tgl = summary.get("total_gain_loss")
        if tgl is not None and not np.isnan(tgl):
            print(f"  Total Gain/Loss: {tgl:,.2f}")
        trp = summary.get("total_return_pct")
        if trp is not None and not np.isnan(trp):
            print(f"  Total Return: {trp:.2%}")

        print("\n  Market Breakdown:")
        for market, stats in summary.get("market_breakdown", {}).items():
            print(f"    {market}: {stats['count']} stocks, "
                  f"cost={stats['cost']:,.0f}, "
                  f"value={stats['market_value']:,.0f}")

        # Export
        output_file = "portfolio_report.xlsx"
        export_portfolio(holdings, summary, output_file)
    else:
        # Demo mode
        print("Portfolio Manager - Demo Mode")
        print("-" * 40)
        print("Usage: python portfolio_manager.py <portfolio.xlsx> [sheet_name]")
        print("\nExpected Excel columns:")
        print("  code (ticker), shares, buy_price, market (optional)")
        print("\nExample:")
        print("  code    shares  buy_price  market")
        print("  AAPL    100     150.00     US")
        print("  2330.TW 1000    500.00     TW")
        print("  9988.HK 200     80.00      HK")

        # Create a sample DataFrame for demo
        demo = pd.DataFrame({
            "code": ["AAPL", "MSFT", "GOOGL"],
            "shares": [100, 50, 30],
            "buy_price": [150.0, 300.0, 2800.0],
            "market": ["US", "US", "US"],
        })
        demo["current_price"] = np.nan
        demo["market_value"] = np.nan
        demo["cost_value"] = demo["shares"] * demo["buy_price"]
        demo["gain_loss"] = np.nan
        demo["gain_loss_pct"] = np.nan
        demo["expected_return"] = np.nan

        print("\nUpdating demo portfolio prices...")
        demo = update_prices(demo)

        summary = generate_portfolio_summary(demo)
        print(f"\n  Demo Portfolio: {summary['total_holdings']} holdings")
        tc = summary.get("total_cost")
        if tc:
            print(f"  Total Cost: ${tc:,.2f}")
        tmv = summary.get("total_market_value")
        if tmv and not np.isnan(tmv):
            print(f"  Total Market Value: ${tmv:,.2f}")


if __name__ == "__main__":
    main()
