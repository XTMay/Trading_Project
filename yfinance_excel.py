"""
yfinance Excel 集成工具
======================
每个功能封装为独立 function，可从 Excel VBA 通过命令行调用。
结果直接写入指定 Excel 文件的指定 Sheet / Cell。

安装依赖:
    source venv/bin/activate
    pip install yfinance openpyxl

命令行用法:
    python yfinance_excel.py info        AAPL  output.xlsx  Sheet1  A1
    python yfinance_excel.py history     AAPL  output.xlsx  Sheet1  A1  --period 1mo --interval 1d
    python yfinance_excel.py history     AAPL  output.xlsx  Sheet1  A1  --start 2025-01-01 --end 2025-06-30
    python yfinance_excel.py financials  AAPL  output.xlsx  Sheet1  A1  --report income
    python yfinance_excel.py dividends   AAPL  output.xlsx  Sheet1  A1  --rows 20
    python yfinance_excel.py holders     AAPL  output.xlsx  Sheet1  A1
    python yfinance_excel.py recommend   AAPL  output.xlsx  Sheet1  A1
    python yfinance_excel.py download    AAPL,MSFT,GOOGL  output.xlsx  Sheet1  A1  --period 5d
    python yfinance_excel.py options     AAPL  output.xlsx  Sheet1  A1  --type calls --expiry 0
"""

import sys
import os
import re
import argparse

import yfinance as yf
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string


# ============================================================
# 工具函数
# ============================================================

def _open_workbook(wb_path):
    """打开已有的 workbook，如果不存在则新建。"""
    if os.path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    wb = openpyxl.Workbook()
    wb.save(wb_path)
    return wb


def _get_or_create_sheet(wb, sheet_name):
    """获取指定 sheet，不存在则新建。"""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)


def _parse_cell(cell_ref):
    """解析 cell 引用 (如 'B3') 为 (row, col) 整数。"""
    m = re.match(r'^([A-Z]+)(\d+)$', cell_ref.upper())
    if not m:
        raise ValueError(f"无效的 cell 引用: {cell_ref}")
    col = column_index_from_string(m.group(1))
    row = int(m.group(2))
    return row, col


def _write_value(ws, row, col, value):
    """将值写入 cell，自动转换 numpy/pandas 类型。"""
    if isinstance(value, (np.integer,)):
        value = int(value)
    elif isinstance(value, (np.floating,)):
        value = float(value)
    elif isinstance(value, (np.bool_,)):
        value = bool(value)
    elif isinstance(value, pd.Timestamp):
        value = value.strftime("%Y-%m-%d")
    elif value is None or (isinstance(value, float) and np.isnan(value)):
        value = ""
    ws.cell(row=row, column=col, value=value)


def _write_dataframe(ws, df, start_row, start_col, write_index=True, index_label=""):
    """将 DataFrame 写入 sheet，包含表头。返回最后写入的行号。"""
    r = start_row
    c = start_col

    if write_index:
        ws.cell(row=r, column=c, value=index_label or df.index.name or "")
        for j, col_name in enumerate(df.columns):
            ws.cell(row=r, column=c + 1 + j, value=str(col_name))
    else:
        for j, col_name in enumerate(df.columns):
            ws.cell(row=r, column=c + j, value=str(col_name))

    for i in range(len(df)):
        r += 1
        offset = 0
        if write_index:
            _write_value(ws, r, c, str(df.index[i]))
            offset = 1
        for j in range(len(df.columns)):
            _write_value(ws, r, c + offset + j, df.iloc[i, j])

    return r


def _save_and_report(wb, wb_path, sheet_name, cell_ref, description):
    """保存 workbook 并打印结果摘要。"""
    wb.save(wb_path)
    print(f"[OK] {description}")
    print(f"     文件: {wb_path}")
    print(f"     位置: {sheet_name}!{cell_ref}")


# ============================================================
# 功能 1: 股票基本信息
# ============================================================

def get_stock_info(symbol, wb_path, sheet_name, cell_ref):
    """
    获取股票基本信息 (公司名、行业、市值、PE、股息率等)。
    纵向写入: A列=字段名, B列=值。
    """
    ticker = yf.Ticker(symbol)
    info = ticker.info

    fields = [
        ("代码",        symbol),
        ("公司名称",    info.get("shortName", "")),
        ("行业",        info.get("industry", "")),
        ("板块",        info.get("sector", "")),
        ("国家",        info.get("country", "")),
        ("市值",        info.get("marketCap", "")),
        ("企业价值",    info.get("enterpriseValue", "")),
        ("PE (TTM)",    info.get("trailingPE", "")),
        ("PE (Forward)",info.get("forwardPE", "")),
        ("PB",          info.get("priceToBook", "")),
        ("EPS (TTM)",   info.get("trailingEps", "")),
        ("股息率",      info.get("dividendYield", "")),
        ("每股股息",    info.get("dividendRate", "")),
        ("52周最高",    info.get("fiftyTwoWeekHigh", "")),
        ("52周最低",    info.get("fiftyTwoWeekLow", "")),
        ("Beta",        info.get("beta", "")),
        ("平均成交量",  info.get("averageVolume", "")),
        ("最新价格",    info.get("currentPrice", "")),
    ]

    wb = _open_workbook(wb_path)
    ws = _get_or_create_sheet(wb, sheet_name)
    start_row, start_col = _parse_cell(cell_ref)

    ws.cell(row=start_row, column=start_col, value="字段")
    ws.cell(row=start_row, column=start_col + 1, value="值")
    for i, (label, val) in enumerate(fields):
        ws.cell(row=start_row + 1 + i, column=start_col, value=label)
        _write_value(ws, start_row + 1 + i, start_col + 1, val)

    _save_and_report(wb, wb_path, sheet_name, cell_ref,
                     f"{symbol} 基本信息 ({len(fields)} 项)")


# ============================================================
# 功能 2: 历史价格
# ============================================================

def get_history(symbol, wb_path, sheet_name, cell_ref,
                period="1mo", interval="1d", start=None, end=None):
    """
    获取历史价格数据。
    支持 period 模式 (如 1mo) 或 start/end 日期范围模式。
    """
    ticker = yf.Ticker(symbol)
    if start and end:
        hist = ticker.history(start=start, end=end, interval=interval)
    else:
        hist = ticker.history(period=period, interval=interval)

    if hist.empty:
        print(f"[WARN] {symbol} 没有历史数据")
        return

    hist.index = hist.index.strftime("%Y-%m-%d")

    wb = _open_workbook(wb_path)
    ws = _get_or_create_sheet(wb, sheet_name)
    start_row, start_col = _parse_cell(cell_ref)

    _write_dataframe(ws, hist, start_row, start_col, write_index=True, index_label="Date")
    _save_and_report(wb, wb_path, sheet_name, cell_ref,
                     f"{symbol} 历史价格 ({len(hist)} 行)")


# ============================================================
# 功能 3: 财务报表
# ============================================================

def get_financials(symbol, wb_path, sheet_name, cell_ref,
                   report="income", quarterly=False):
    """
    获取财务报表。
    report: 'income' (损益表), 'balance' (资产负债表), 'cashflow' (现金流量表)
    quarterly: True 获取季度数据
    """
    ticker = yf.Ticker(symbol)
    report_map = {
        "income":   ticker.quarterly_financials if quarterly else ticker.financials,
        "balance":  ticker.quarterly_balance_sheet if quarterly else ticker.balance_sheet,
        "cashflow": ticker.quarterly_cashflow if quarterly else ticker.cashflow,
    }

    df = report_map.get(report)
    if df is None or df.empty:
        print(f"[WARN] {symbol} 没有 {report} 报表数据")
        return

    df.columns = [c.strftime("%Y-%m-%d") if hasattr(c, "strftime") else str(c)
                  for c in df.columns]

    wb = _open_workbook(wb_path)
    ws = _get_or_create_sheet(wb, sheet_name)
    start_row, start_col = _parse_cell(cell_ref)

    report_names = {"income": "损益表", "balance": "资产负债表", "cashflow": "现金流量表"}
    period_label = "季度" if quarterly else "年度"

    _write_dataframe(ws, df, start_row, start_col, write_index=True, index_label="科目")
    _save_and_report(wb, wb_path, sheet_name, cell_ref,
                     f"{symbol} {period_label}{report_names.get(report)} ({len(df)} 行)")


# ============================================================
# 功能 4: 股息与拆股历史
# ============================================================

def get_dividends(symbol, wb_path, sheet_name, cell_ref, rows=20):
    """
    获取股息 + 拆股历史。先写股息，空 2 行后写拆股。
    """
    ticker = yf.Ticker(symbol)

    wb = _open_workbook(wb_path)
    ws = _get_or_create_sheet(wb, sheet_name)
    start_row, start_col = _parse_cell(cell_ref)

    # 股息
    divs = ticker.dividends.tail(rows)
    ws.cell(row=start_row, column=start_col, value="日期")
    ws.cell(row=start_row, column=start_col + 1, value="股息")
    for i, (date, val) in enumerate(divs.items()):
        ws.cell(row=start_row + 1 + i, column=start_col, value=str(date)[:10])
        _write_value(ws, start_row + 1 + i, start_col + 1, float(val))

    # 拆股 (空 2 行)
    gap_row = start_row + 1 + len(divs) + 2
    splits = ticker.splits
    ws.cell(row=gap_row, column=start_col, value="日期")
    ws.cell(row=gap_row, column=start_col + 1, value="拆股比例")
    for i, (date, val) in enumerate(splits.items()):
        ws.cell(row=gap_row + 1 + i, column=start_col, value=str(date)[:10])
        _write_value(ws, gap_row + 1 + i, start_col + 1, float(val))

    _save_and_report(wb, wb_path, sheet_name, cell_ref,
                     f"{symbol} 股息 {len(divs)} 笔, 拆股 {len(splits)} 笔")


# ============================================================
# 功能 5: 持有人信息
# ============================================================

def get_holders(symbol, wb_path, sheet_name, cell_ref):
    """
    获取主要持有人 + 机构持有人 (前10)。
    """
    ticker = yf.Ticker(symbol)

    wb = _open_workbook(wb_path)
    ws = _get_or_create_sheet(wb, sheet_name)
    start_row, start_col = _parse_cell(cell_ref)
    r = start_row

    # 主要持有人
    mh = ticker.major_holders
    if mh is not None and not mh.empty:
        ws.cell(row=r, column=start_col, value="主要持有人")
        r += 1
        for i in range(len(mh)):
            _write_value(ws, r + i, start_col, mh.iloc[i, 0])
            if len(mh.columns) > 1:
                _write_value(ws, r + i, start_col + 1, mh.iloc[i, 1])
        r += len(mh) + 1

    # 机构持有人
    ih = ticker.institutional_holders
    if ih is not None and not ih.empty:
        ws.cell(row=r, column=start_col, value="机构持有人 (前10)")
        r += 1
        _write_dataframe(ws, ih.head(10), r, start_col, write_index=False)

    _save_and_report(wb, wb_path, sheet_name, cell_ref, f"{symbol} 持有人信息")


# ============================================================
# 功能 6: 分析师建议
# ============================================================

def get_recommendations(symbol, wb_path, sheet_name, cell_ref):
    """获取分析师建议。"""
    ticker = yf.Ticker(symbol)
    rec = ticker.recommendations

    if rec is None or rec.empty:
        print(f"[WARN] {symbol} 没有分析师建议数据")
        return

    wb = _open_workbook(wb_path)
    ws = _get_or_create_sheet(wb, sheet_name)
    start_row, start_col = _parse_cell(cell_ref)

    _write_dataframe(ws, rec, start_row, start_col, write_index=False)
    _save_and_report(wb, wb_path, sheet_name, cell_ref,
                     f"{symbol} 分析师建议 ({len(rec)} 行)")


# ============================================================
# 功能 7: 批量下载多只股票
# ============================================================

def download_multiple(symbols_str, wb_path, sheet_name, cell_ref,
                      period="1mo", interval="1d"):
    """
    批量下载多只股票。symbols_str: 逗号分隔 (如 "AAPL,MSFT,GOOGL")。
    每只股票水平排列，之间空 1 列。
    """
    symbols = [s.strip() for s in symbols_str.split(",")]

    wb = _open_workbook(wb_path)
    ws = _get_or_create_sheet(wb, sheet_name)
    start_row, start_col = _parse_cell(cell_ref)

    col_offset = 0
    for sym in symbols:
        ticker = yf.Ticker(sym)
        hist = ticker.history(period=period, interval=interval)
        if hist.empty:
            continue
        hist.index = hist.index.strftime("%Y-%m-%d")

        ws.cell(row=start_row, column=start_col + col_offset, value=sym)
        _write_dataframe(ws, hist, start_row + 1, start_col + col_offset,
                         write_index=True, index_label="Date")
        col_offset += 1 + len(hist.columns) + 1

    _save_and_report(wb, wb_path, sheet_name, cell_ref,
                     f"批量下载 {len(symbols)} 只股票")


# ============================================================
# 功能 8: 期权数据
# ============================================================

def get_options(symbol, wb_path, sheet_name, cell_ref,
                opt_type="calls", expiry_index=0):
    """
    获取期权数据。
    opt_type: 'calls' 或 'puts'
    expiry_index: 到期日索引 (0 = 最近的到期日)
    """
    ticker = yf.Ticker(symbol)
    expirations = ticker.options

    if not expirations:
        print(f"[WARN] {symbol} 没有期权数据")
        return

    idx = min(int(expiry_index), len(expirations) - 1)
    exp_date = expirations[idx]
    chain = ticker.option_chain(exp_date)
    df = chain.calls if opt_type == "calls" else chain.puts

    wb = _open_workbook(wb_path)
    ws = _get_or_create_sheet(wb, sheet_name)
    start_row, start_col = _parse_cell(cell_ref)

    ws.cell(row=start_row, column=start_col,
            value=f"{symbol} {opt_type.upper()} - 到期日: {exp_date}")
    _write_dataframe(ws, df, start_row + 1, start_col, write_index=False)
    _save_and_report(wb, wb_path, sheet_name, cell_ref,
                     f"{symbol} {opt_type} 期权 ({exp_date}, {len(df)} 行)")


# ============================================================
# 命令行入口 (供 VBA Shell 调用)
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="yfinance Excel 集成工具")
    subparsers = parser.add_subparsers(dest="command", help="可用命令")

    # info
    p = subparsers.add_parser("info", help="股票基本信息")
    p.add_argument("symbol");  p.add_argument("workbook")
    p.add_argument("sheet");   p.add_argument("cell")

    # history
    p = subparsers.add_parser("history", help="历史价格")
    p.add_argument("symbol");  p.add_argument("workbook")
    p.add_argument("sheet");   p.add_argument("cell")
    p.add_argument("--period", default="1mo")
    p.add_argument("--interval", default="1d")
    p.add_argument("--start", default=None)
    p.add_argument("--end", default=None)

    # financials
    p = subparsers.add_parser("financials", help="财务报表")
    p.add_argument("symbol");  p.add_argument("workbook")
    p.add_argument("sheet");   p.add_argument("cell")
    p.add_argument("--report", default="income", choices=["income", "balance", "cashflow"])
    p.add_argument("--quarterly", action="store_true")

    # dividends
    p = subparsers.add_parser("dividends", help="股息与拆股")
    p.add_argument("symbol");  p.add_argument("workbook")
    p.add_argument("sheet");   p.add_argument("cell")
    p.add_argument("--rows", type=int, default=20)

    # holders
    p = subparsers.add_parser("holders", help="持有人信息")
    p.add_argument("symbol");  p.add_argument("workbook")
    p.add_argument("sheet");   p.add_argument("cell")

    # recommend
    p = subparsers.add_parser("recommend", help="分析师建议")
    p.add_argument("symbol");  p.add_argument("workbook")
    p.add_argument("sheet");   p.add_argument("cell")

    # download
    p = subparsers.add_parser("download", help="批量下载")
    p.add_argument("symbols", help="逗号分隔: AAPL,MSFT,GOOGL")
    p.add_argument("workbook"); p.add_argument("sheet"); p.add_argument("cell")
    p.add_argument("--period", default="1mo")
    p.add_argument("--interval", default="1d")

    # options
    p = subparsers.add_parser("options", help="期权数据")
    p.add_argument("symbol");  p.add_argument("workbook")
    p.add_argument("sheet");   p.add_argument("cell")
    p.add_argument("--type", dest="opt_type", default="calls", choices=["calls", "puts"])
    p.add_argument("--expiry", type=int, default=0)

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        sys.exit(1)

    dispatch = {
        "info":       lambda: get_stock_info(args.symbol, args.workbook, args.sheet, args.cell),
        "history":    lambda: get_history(args.symbol, args.workbook, args.sheet, args.cell,
                                          args.period, args.interval, args.start, args.end),
        "financials": lambda: get_financials(args.symbol, args.workbook, args.sheet, args.cell,
                                             args.report, args.quarterly),
        "dividends":  lambda: get_dividends(args.symbol, args.workbook, args.sheet, args.cell, args.rows),
        "holders":    lambda: get_holders(args.symbol, args.workbook, args.sheet, args.cell),
        "recommend":  lambda: get_recommendations(args.symbol, args.workbook, args.sheet, args.cell),
        "download":   lambda: download_multiple(args.symbols, args.workbook, args.sheet, args.cell,
                                                args.period, args.interval),
        "options":    lambda: get_options(args.symbol, args.workbook, args.sheet, args.cell,
                                          args.opt_type, args.expiry),
    }
    dispatch[args.command]()


if __name__ == "__main__":
    main()
