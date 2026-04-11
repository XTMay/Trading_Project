"""
generate_report_summary.py
Unified script that fetches all financial data for one ticker via yfinance
and outputs report_summary.xlsx matching the 盈再表 "美股" sheet layout.

Usage:
    python generate_report_summary.py              # uses default ticker
    python generate_report_summary.py AAPL          # specify ticker
    python generate_report_summary.py 9022.T JPY    # specify ticker + currency
"""

import sys
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook

# ========= Configuration =========
DEFAULT_TICKER = "AAPL"
DEFAULT_CURRENCY = None  # Auto-detect from financialCurrency; set e.g. "JPY" to override

OUTPUT_FILE = "report_summary.xlsx"
SHEET_NAME = "美股"


# =====================================================================
# Data Fetching Functions
# =====================================================================

def fetch_quarterly_income(ticker):
    """Fetch quarterly income statement: last 5 quarters + TTM (sum of last 4Q)."""
    df = ticker.quarterly_financials
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = pd.to_datetime(df.columns)
    df = df.sort_index(axis=1)

    # Select last 5 quarters
    last_q = df.iloc[:, -5:] if df.shape[1] >= 5 else df.copy()
    last_q.columns = last_q.columns.strftime("%Y-%m")

    # TTM = sum of last 4 quarters
    if df.shape[1] >= 4:
        last_4q = df.iloc[:, -4:]
        ttm = last_4q.apply(pd.to_numeric, errors="coerce").sum(axis=1).to_frame(name="TTM")
    else:
        ttm = df.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
        ttm.columns = ["TTM"]

    combined = last_q.merge(ttm, left_index=True, right_index=True, how="left")

    # Unusual items
    keywords = ["unusual", "special", "restructuring", "non recurring"]
    unusual_mask = combined.index.str.lower().str.contains("|".join(keywords))
    unusual = combined.loc[unusual_mask] if unusual_mask.any() else pd.DataFrame()

    return _sanitize(combined), unusual


def fetch_quarterly_balance_sheet(ticker):
    """Fetch quarterly balance sheet: last 5 quarters + TTM (latest snapshot)."""
    df = ticker.quarterly_balance_sheet
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = pd.to_datetime(df.columns)
    df = df.sort_index(axis=1)

    last_q = df.iloc[:, -5:] if df.shape[1] >= 5 else df.copy()
    last_q.columns = last_q.columns.strftime("%Y-%m")

    # TTM for BS = latest quarter snapshot
    ttm = df.iloc[:, -1:].copy()
    ttm.columns = ["TTM"]

    combined = last_q.merge(ttm, left_index=True, right_index=True, how="left")
    return _sanitize(combined)


def fetch_annual_income(ticker):
    """Fetch annual income statement: last 5 years + TTM."""
    df = ticker.financials
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = pd.to_datetime(df.columns)
    df = df.sort_index(axis=1)

    # Exclude current year
    current_year = datetime.today().year
    df = df.loc[:, df.columns.year < current_year]
    if df.empty:
        return pd.DataFrame()

    last_y = df.iloc[:, -5:] if df.shape[1] >= 5 else df.copy()
    last_y.columns = last_y.columns.strftime("%Y")

    # TTM = most recent year
    ttm = df.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
    ttm.columns = ["TTM"]

    combined = last_y.merge(ttm, left_index=True, right_index=True, how="left")
    return _sanitize(combined)


def fetch_annual_balance_sheet(ticker):
    """Fetch annual balance sheet: last 5 years + TTM (latest snapshot)."""
    df = ticker.balance_sheet
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = pd.to_datetime(df.columns)
    df = df.sort_index(axis=1)

    last_y = df.iloc[:, -5:] if df.shape[1] >= 5 else df.copy()
    last_y.columns = last_y.columns.strftime("%Y")

    ttm = df.iloc[:, -1:].copy()
    ttm.columns = ["TTM"]

    combined = last_y.merge(ttm, left_index=True, right_index=True, how="left")
    return _sanitize(combined)


def fetch_annual_cashflow(ticker):
    """Fetch annual cash flow: last 5 years + TTM (4Q sum or fallback)."""
    cf_y = ticker.cashflow
    if cf_y is None or cf_y.empty:
        return pd.DataFrame()
    cf_y.columns = pd.to_datetime(cf_y.columns)
    cf_y = cf_y.sort_index(axis=1)

    last_y = cf_y.iloc[:, -5:] if cf_y.shape[1] >= 5 else cf_y.copy()
    last_y.columns = last_y.columns.strftime("%Y")

    # TTM from quarterly cashflow
    try:
        cf_q = ticker.quarterly_cashflow
        if cf_q is not None and not cf_q.empty:
            cf_q.columns = pd.to_datetime(cf_q.columns)
            cf_q = cf_q.sort_index(axis=1)
            if cf_q.shape[1] >= 4:
                ttm = cf_q.iloc[:, -4:].apply(pd.to_numeric, errors="coerce").sum(axis=1).to_frame(name="TTM")
            else:
                ttm = cf_q.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
                ttm.columns = ["TTM"]
        else:
            ttm = cf_y.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
            ttm.columns = ["TTM"]
    except Exception:
        ttm = cf_y.iloc[:, -1:].apply(pd.to_numeric, errors="coerce").copy()
        ttm.columns = ["TTM"]

    combined = last_y.merge(ttm, left_index=True, right_index=True, how="left")
    return _sanitize(combined)


def fetch_company_profile(ticker):
    """Fetch company profile from ticker.info."""
    try:
        info = ticker.info
    except Exception:
        return {}
    if not info:
        return {}
    return {
        "longName": info.get("longName", ""),
        "symbol": info.get("symbol", ""),
        "exchange": info.get("exchange", ""),
        "sector": info.get("sector", ""),
        "industry": info.get("industry", ""),
        "country": info.get("country", ""),
        "currency": info.get("currency", ""),
        "financialCurrency": info.get("financialCurrency", "USD"),
        "currentPrice": info.get("currentPrice"),
        "marketCap": info.get("marketCap"),
        "enterpriseValue": info.get("enterpriseValue"),
        "trailingPE": info.get("trailingPE"),
        "forwardPE": info.get("forwardPE"),
        "trailingEps": info.get("trailingEps"),
        "dividendYield": info.get("dividendYield"),
        "payoutRatio": info.get("payoutRatio"),
        "beta": info.get("beta"),
        "fiftyTwoWeekHigh": info.get("fiftyTwoWeekHigh"),
        "fiftyTwoWeekLow": info.get("fiftyTwoWeekLow"),
        "longBusinessSummary": info.get("longBusinessSummary", ""),
        "website": info.get("website", ""),
        "totalRevenue": info.get("totalRevenue"),
        "netIncomeToCommon": info.get("netIncomeToCommon"),
    }


def fetch_share_capital(ticker):
    """Fetch share capital data from ticker.info."""
    try:
        info = ticker.info
    except Exception:
        return {}
    if not info:
        return {}
    return {
        "Shares Outstanding": info.get("sharesOutstanding"),
        "Float Shares": info.get("floatShares"),
        "Implied Shares Outstanding": info.get("impliedSharesOutstanding"),
        "Held by Insiders (%)": info.get("heldPercentInsiders"),
        "Held by Institutions (%)": info.get("heldPercentInstitutions"),
        "Short Shares": info.get("sharesShort"),
        "Short Prior Month": info.get("sharesShortPriorMonth"),
        "Short % of Shares Outstanding": info.get("shortPercentOfFloat"),
    }


def fetch_eps_earnings(ticker):
    """Fetch EPS / earnings data."""
    result = {"annual": {}, "quarterly": {}, "ttm_eps": None}

    # Annual net income from income statement (replaces deprecated ticker.earnings)
    try:
        is_y = ticker.income_stmt
        if is_y is not None and not is_y.empty:
            is_y.columns = pd.to_datetime(is_y.columns)
            is_y = is_y.sort_index(axis=1)
            for label in ["Net Income", "Net Income Common Stockholders"]:
                if label in is_y.index:
                    for col in is_y.columns:
                        val = is_y.loc[label, col]
                        if pd.notna(val):
                            result["annual"][col.strftime("%Y")] = val
                    break
    except Exception:
        pass

    # Quarterly net income
    try:
        is_q = ticker.quarterly_income_stmt
        if is_q is not None and not is_q.empty:
            is_q.columns = pd.to_datetime(is_q.columns)
            is_q = is_q.sort_index(axis=1)
            for label in ["Net Income", "Net Income Common Stockholders"]:
                if label in is_q.index:
                    for col in is_q.columns:
                        val = is_q.loc[label, col]
                        if pd.notna(val):
                            result["quarterly"][col.strftime("%Y-%m-%d")] = val
                    break
    except Exception:
        pass

    # TTM EPS from info
    try:
        info = ticker.info
        result["ttm_eps"] = info.get("trailingEps")
    except Exception:
        pass
    return result


def fetch_historical_prices_adj(ticker, start_date="2014-01-01"):
    """Fetch monthly historical prices with Adj Close and corporate actions
    (based on Historical_stock_price_adj.py logic)."""
    try:
        end_date = (datetime.today() + timedelta(days=1)).strftime("%Y-%m-%d")
        df = ticker.history(start=start_date, end=end_date, auto_adjust=False)
    except Exception:
        return pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame()

    # Monthly resampling (month start)
    monthly_df = df.resample('MS').agg({
        'Open': 'first',
        'High': 'max',
        'Low': 'min',
        'Close': 'last',
        'Adj Close': 'last',
        'Volume': 'sum'
    })

    # Corporate Actions
    try:
        actions = ticker.actions
        if actions is not None and not actions.empty:
            actions = actions.loc[start_date:]
            actions = actions[
                (actions['Dividends'] != 0) |
                (actions['Stock Splits'] != 0)
            ]
        else:
            actions = pd.DataFrame()
    except Exception:
        actions = pd.DataFrame()

    # Merge price data with corporate actions
    combined = pd.concat([monthly_df, actions], axis=0)

    # Ensure Dividends / Stock Splits columns exist
    for col in ['Dividends', 'Stock Splits']:
        if col not in combined.columns:
            combined[col] = 0
        else:
            combined[col] = combined[col].fillna(0)

    # Entry_Type label
    combined['Entry_Type'] = "Month_Start"
    if not actions.empty:
        combined.loc[actions.index, 'Entry_Type'] = "Corporate_Action"

    # Sort descending (newest first)
    combined = combined.sort_index(ascending=False)

    return combined


def fetch_exchange_rate(currency_code):
    """Fetch USD/XXX exchange rate."""
    if not currency_code or currency_code.upper() == "USD":
        return {"rate": 1.0, "pair": "USD/USD"}
    symbol = f"USD{currency_code.upper()}=X"
    try:
        t = yf.Ticker(symbol)
        data = t.history(period="5d")
        if data is not None and not data.empty:
            latest = data.iloc[-1]
            return {
                "pair": f"USD/{currency_code.upper()}",
                "rate": latest["Close"],
                "open": latest.get("Open"),
                "high": latest.get("High"),
                "low": latest.get("Low"),
            }
    except Exception:
        pass
    return {"rate": 1.0, "pair": f"USD/{currency_code.upper()}"}


# =====================================================================
# Helper
# =====================================================================

def _sanitize(df):
    """Clean DataFrame for Excel output."""
    df = df.copy()
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.replace([np.inf, -np.inf], np.nan)
    return df


def _safe_val(val):
    """Return a safe value for writing to Excel (no NaN/inf, no formula injection)."""
    if val is None:
        return ""
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return ""
    # Prevent Excel formula injection from data values
    if isinstance(val, str) and len(val) > 0 and val[0] in ("=", "+", "-", "@"):
        return "'" + val
    return val


# =====================================================================
# Excel Writing Functions
# =====================================================================

def write_front_page(ws, profile, eps_data, share_cap, forex, net_income_ttm):
    """Write front-page cells: A1, A2, I1, F15, K3, Y16, Y23, Y24, W25, E9, K10-K12."""
    # A1: {Company Name} {Ticker} ({Exchange}) : {Sector}*{Industry}
    company_line = (
        f"{profile.get('longName', '')} {profile.get('symbol', '')} "
        f"({profile.get('exchange', '')}) : "
        f"{profile.get('sector', '')}*{profile.get('industry', '')}"
    )
    ws.cell(row=1, column=1, value=company_line)

    # A2: Ticker symbol
    ws.cell(row=2, column=1, value=profile.get("symbol", ""))

    # I1: Today's date
    ws.cell(row=1, column=9, value=datetime.today().strftime("%Y-%m-%d"))

    # F15: Exchange rate
    ws.cell(row=15, column=6, value=_safe_val(forex.get("rate", 1.0)))

    # K3: Current stock price
    ws.cell(row=3, column=11, value=_safe_val(profile.get("currentPrice")))

    # Y16: Company description (longBusinessSummary)
    ws.cell(row=16, column=25, value=_safe_val(profile.get("longBusinessSummary", "")))

    # Y23: Market Cap (in millions)
    mc = profile.get("marketCap")
    if mc is not None:
        ws.cell(row=23, column=25, value=round(mc / 1_000_000, 2))
    else:
        ws.cell(row=23, column=25, value="")

    # Y24: Stock price
    ws.cell(row=24, column=25, value=_safe_val(profile.get("currentPrice")))

    # W25: Financial currency code
    ws.cell(row=25, column=23, value=_safe_val(profile.get("financialCurrency", "USD")))

    # E9: Net Income TTM (millions)
    if net_income_ttm is not None:
        ws.cell(row=9, column=5, value=round(net_income_ttm / 1_000_000, 2))
    else:
        ws.cell(row=9, column=5, value="")

    # K10: 12 (constant — months)
    ws.cell(row=10, column=11, value=12)

    # K11: Payout ratio
    ws.cell(row=11, column=11, value=_safe_val(profile.get("payoutRatio")))

    # K12: 0.25 (dividend tax default)
    ws.cell(row=12, column=11, value=0.25)


def write_financial_table(ws, df, start_col, section_title, section_num):
    """
    Write a financial table (IS/BS/CFS) to the worksheet.
    Row 1: section header
    Row 2: "Item" + date columns
    Row 3+: data rows
    """
    if df.empty:
        ws.cell(row=1, column=start_col, value=f"{section_num} / 9 {section_title}")
        return

    # Row 1: Section header
    ws.cell(row=1, column=start_col, value=f"{section_num} / 9 {section_title}")

    # Row 2: Column headers
    ws.cell(row=2, column=start_col, value="Item")
    for ci, col_name in enumerate(df.columns, start=1):
        ws.cell(row=2, column=start_col + ci, value=col_name)

    # Row 3+: Data
    for ri, (item_name, row_data) in enumerate(df.iterrows(), start=3):
        ws.cell(row=ri, column=start_col, value=item_name)
        for ci, col_name in enumerate(df.columns, start=1):
            val = row_data[col_name]
            ws.cell(row=ri, column=start_col + ci, value=_safe_val(val))


def write_unusual_items(ws, unusual_df, start_col, data_row_count):
    """Append unusual items below the main IS quarterly table."""
    if unusual_df.empty:
        return
    # Place separator + unusual items after the main data
    sep_row = 3 + data_row_count + 1
    ws.cell(row=sep_row, column=start_col, value="[Unusual / Special Items]")

    for ri, (item_name, row_data) in enumerate(unusual_df.iterrows(), start=sep_row + 1):
        ws.cell(row=ri, column=start_col, value=item_name)
        for ci, col_name in enumerate(unusual_df.columns, start=1):
            val = row_data[col_name]
            ws.cell(row=ri, column=start_col + ci, value=_safe_val(val))


def write_price_history(ws, combined_df, start_col):
    """
    Write historical price section with adjusted close and corporate actions.
    Row 1 (BO1): section header
    Row 4 (BO4:BT4+): column headers
    Row 5+: combined monthly price + corporate action data (newest first)
    """
    ws.cell(row=1, column=start_col, value="5 / 9 Historical stock price (adj)")

    if combined_df.empty:
        return

    # Row 4: Column headers — BO4:BT4 = Date, Open, High, Low, Close, Adj Close
    # BU4+: Volume, Dividends, Stock Splits
    headers = ["Date", "Open", "High", "Low", "Close", "Adj Close",
               "Volume", "Dividends", "Stock Splits"]
    for ci, h in enumerate(headers):
        ws.cell(row=4, column=start_col + ci, value=h)

    # Row 5+: Data rows (already sorted descending)
    data_cols = ["Open", "High", "Low", "Close", "Adj Close",
                 "Volume", "Dividends", "Stock Splits"]
    row_offset = 5
    for ri, (dt_idx, row_data) in enumerate(combined_df.iterrows()):
        r = row_offset + ri
        # Date column
        if isinstance(dt_idx, str):
            ws.cell(row=r, column=start_col, value=dt_idx)
        else:
            ws.cell(row=r, column=start_col, value=dt_idx.strftime("%Y/%m/%d"))
        # Data columns
        for ci, col in enumerate(data_cols, start=1):
            if col in combined_df.columns:
                ws.cell(row=r, column=start_col + ci, value=_safe_val(row_data[col]))


def write_profile_section(ws, profile, start_col):
    """Write company profile raw data section (6/9)."""
    ws.cell(row=1, column=start_col, value="6 / 9 Company Profile")

    items = [
        ("Company Name", profile.get("longName")),
        ("Ticker", profile.get("symbol")),
        ("Exchange", profile.get("exchange")),
        ("Sector", profile.get("sector")),
        ("Industry", profile.get("industry")),
        ("Country", profile.get("country")),
        ("Currency", profile.get("currency")),
        ("Financial Currency", profile.get("financialCurrency")),
        ("Current Price", profile.get("currentPrice")),
        ("Market Cap", profile.get("marketCap")),
        ("Enterprise Value", profile.get("enterpriseValue")),
        ("Trailing PE", profile.get("trailingPE")),
        ("Forward PE", profile.get("forwardPE")),
        ("EPS (TTM)", profile.get("trailingEps")),
        ("Dividend Yield", profile.get("dividendYield")),
        ("Payout Ratio", profile.get("payoutRatio")),
        ("Beta", profile.get("beta")),
        ("52 Week High", profile.get("fiftyTwoWeekHigh")),
        ("52 Week Low", profile.get("fiftyTwoWeekLow")),
        ("Revenue (TTM)", profile.get("totalRevenue")),
        ("Net Income (TTM)", profile.get("netIncomeToCommon")),
        ("Website", profile.get("website")),
        ("Description", profile.get("longBusinessSummary")),
    ]
    for ri, (item, value) in enumerate(items, start=2):
        ws.cell(row=ri, column=start_col, value=item)
        ws.cell(row=ri, column=start_col + 1, value=_safe_val(value))


def write_forex_section(ws, forex, start_col):
    """Write exchange rate section (7/9)."""
    ws.cell(row=1, column=start_col, value="7 / 9 Foreign exchange rate")

    items = [
        ("Currency Pair", forex.get("pair", "")),
        ("Exchange Rate (Close)", forex.get("rate")),
        ("Open", forex.get("open")),
        ("High", forex.get("high")),
        ("Low", forex.get("low")),
    ]
    for ri, (item, value) in enumerate(items, start=2):
        ws.cell(row=ri, column=start_col, value=item)
        ws.cell(row=ri, column=start_col + 1, value=_safe_val(value))


def write_market_cap_section(ws, profile, share_cap, eps_data, start_col):
    """Write market cap / valuation section (9/9)."""
    ws.cell(row=1, column=start_col, value="9 / 9 Market capitalization")

    mc = profile.get("marketCap")
    ev = profile.get("enterpriseValue")
    items = [
        ("Market Cap ($M)", round(mc / 1_000_000, 2) if mc else ""),
        ("Enterprise Value ($M)", round(ev / 1_000_000, 2) if ev else ""),
        ("Trailing PE", profile.get("trailingPE")),
        ("Forward PE", profile.get("forwardPE")),
        ("EPS (TTM)", profile.get("trailingEps")),
        ("Dividend Yield", profile.get("dividendYield")),
        ("Payout Ratio", profile.get("payoutRatio")),
        ("Beta", profile.get("beta")),
        ("52 Week High", profile.get("fiftyTwoWeekHigh")),
        ("52 Week Low", profile.get("fiftyTwoWeekLow")),
        ("Revenue (TTM)", profile.get("totalRevenue")),
        ("Net Income (TTM)", profile.get("netIncomeToCommon")),
        ("Shares Outstanding", share_cap.get("Shares Outstanding")),
        ("Float Shares", share_cap.get("Float Shares")),
        ("Implied Shares Outstanding", share_cap.get("Implied Shares Outstanding")),
        ("Held by Insiders (%)", share_cap.get("Held by Insiders (%)")),
        ("Held by Institutions (%)", share_cap.get("Held by Institutions (%)")),
        ("Short Shares", share_cap.get("Short Shares")),
        ("Short Prior Month", share_cap.get("Short Prior Month")),
        ("Short % of Shares Outstanding", share_cap.get("Short % of Shares Outstanding")),
    ]

    # Append EPS earnings data
    if eps_data.get("ttm_eps") is not None:
        items.append(("TTM EPS", eps_data["ttm_eps"]))
    if eps_data.get("annual"):
        items.append(("[Annual Earnings]", ""))
        for year, val in eps_data["annual"].items():
            items.append((f"Earnings {year}", val))
    if eps_data.get("quarterly"):
        items.append(("[Quarterly Earnings]", ""))
        for date_str, val in eps_data["quarterly"].items():
            items.append((f"Earnings {date_str}", val))

    for ri, (item, value) in enumerate(items, start=2):
        ws.cell(row=ri, column=start_col, value=item)
        ws.cell(row=ri, column=start_col + 1, value=_safe_val(value))


# =====================================================================
# Main
# =====================================================================

def main():
    # Parse CLI args
    ticker_symbol = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_TICKER
    currency_override = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_CURRENCY

    print(f"{'='*60}")
    print(f"  Report Summary Generator")
    print(f"  Ticker: {ticker_symbol}")
    print(f"{'='*60}")

    ticker = yf.Ticker(ticker_symbol)

    # ---- Fetch all data ----
    print("1/10 Fetching quarterly income statement...")
    is_q_result = fetch_quarterly_income(ticker)
    if isinstance(is_q_result, tuple):
        is_q, unusual_q = is_q_result
    else:
        is_q, unusual_q = is_q_result, pd.DataFrame()

    print("2/10 Fetching quarterly balance sheet...")
    bs_q = fetch_quarterly_balance_sheet(ticker)

    print("3/10 Fetching annual income statement...")
    is_y = fetch_annual_income(ticker)

    print("4/10 Fetching annual balance sheet...")
    bs_y = fetch_annual_balance_sheet(ticker)

    print("5/10 Fetching annual cash flow statement...")
    cfs_y = fetch_annual_cashflow(ticker)

    print("6/10 Fetching company profile...")
    profile = fetch_company_profile(ticker)

    print("7/10 Fetching share capital data...")
    share_cap = fetch_share_capital(ticker)

    print("8/10 Fetching EPS / earnings data...")
    eps_data = fetch_eps_earnings(ticker)

    print("9/10 Fetching historical prices (adjusted) & actions...")
    combined_prices = fetch_historical_prices_adj(ticker)

    # Determine currency and fetch exchange rate
    fin_currency = currency_override or profile.get("financialCurrency", "USD")
    print(f"10/10 Fetching exchange rate (USD/{fin_currency})...")
    forex = fetch_exchange_rate(fin_currency)

    # ---- Compute Net Income TTM for front page ----
    net_income_ttm = None
    if not is_q.empty:
        for label in ["Net Income", "Net Income Common Stockholders"]:
            if label in is_q.index and "TTM" in is_q.columns:
                val = is_q.loc[label, "TTM"]
                if pd.notna(val):
                    net_income_ttm = val
                    break
    # Fallback to ticker.info netIncomeToCommon
    if net_income_ttm is None and profile.get("netIncomeToCommon") is not None:
        net_income_ttm = profile["netIncomeToCommon"]

    # ---- Column positions (美股 sheet layout) ----
    COL_IS_Q = 31   # AE — 1/9 Quarterly IS
    COL_BS_Q = 40   # AN — 2/9 Quarterly BS
    COL_IS_Y = 49   # AW — 3/9 Annual IS
    COL_BS_Y = 58   # BF — 4/9 Annual BS
    COL_PRICE = 67  # BO — 5/9 Historical Price
    COL_PROF = 76   # BX — 6/9 Company Profile
    COL_FX = 79     # CA — 7/9 Exchange Rate
    COL_CFS = 83    # CE — 8/9 Annual CFS
    COL_MCAP = 92   # CN — 9/9 Market Cap

    # ---- Create workbook ----
    print("\nWriting Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Front page cells
    write_front_page(ws, profile, eps_data, share_cap, forex, net_income_ttm)

    # 1/9 Quarterly Income Statement
    write_financial_table(ws, is_q, COL_IS_Q, "Quarterly Income Statement", 1)
    if not unusual_q.empty:
        write_unusual_items(ws, unusual_q, COL_IS_Q, len(is_q))

    # 2/9 Quarterly Balance Sheet
    write_financial_table(ws, bs_q, COL_BS_Q, "Quarterly Balance Sheet", 2)

    # 3/9 Annual Income Statement
    write_financial_table(ws, is_y, COL_IS_Y, "Annual Income Statement", 3)

    # 4/9 Annual Balance Sheet
    write_financial_table(ws, bs_y, COL_BS_Y, "Annual Balance Sheet", 4)

    # 5/9 Historical Stock Price (adj)
    write_price_history(ws, combined_prices, COL_PRICE)

    # 6/9 Company Profile
    write_profile_section(ws, profile, COL_PROF)

    # 7/9 Exchange Rate
    write_forex_section(ws, forex, COL_FX)

    # 8/9 Annual Cash Flow
    write_financial_table(ws, cfs_y, COL_CFS, "Annual Cash Flow", 8)

    # 9/9 Market Capitalization
    write_market_cap_section(ws, profile, share_cap, eps_data, COL_MCAP)

    # ---- Save ----
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved: {OUTPUT_FILE}")
    print(f"   Sheet: {SHEET_NAME}")
    print(f"   Sections written:")
    print(f"     1/9 Quarterly IS  → col AE ({COL_IS_Q})")
    print(f"     2/9 Quarterly BS  → col AN ({COL_BS_Q})")
    print(f"     3/9 Annual IS     → col AW ({COL_IS_Y})")
    print(f"     4/9 Annual BS     → col BF ({COL_BS_Y})")
    print(f"     5/9 Price History (adj) → col BO ({COL_PRICE}), headers BO4:BW4")
    print(f"     6/9 Profile       → col BX ({COL_PROF})")
    print(f"     7/9 Exchange Rate → col CA ({COL_FX})")
    print(f"     8/9 Annual CFS    → col CE ({COL_CFS})")
    print(f"     9/9 Market Cap    → col CN ({COL_MCAP})")
    print(f"\n   Front-page cells: A1, A2, I1, F15, K3, Y16, Y23, Y24, W25, E9, K10-K12")


if __name__ == "__main__":
    main()
