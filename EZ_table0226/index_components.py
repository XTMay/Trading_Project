"""
index_components.py
Market Index Components Engine for 盈再表.

Implements VBA Module6 (high feasibility items):
  - Fetch S&P 500 components from Wikipedia
  - Fetch TWSE weighted index components
  - Sort by P/E ratio
  - Enrich with yfinance financial metrics

Usage:
    python index_components.py              # fetch S&P 500 components
    python index_components.py sp500        # fetch S&P 500
    python index_components.py twse         # fetch TWSE components
"""

import sys
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook

from generate_report_summary import _safe_val

# =====================================================================
# S&P 500 Components
# =====================================================================


def fetch_sp500_components():
    """Fetch S&P 500 component list from Wikipedia.

    Returns:
        pd.DataFrame with columns: symbol, company, sector, sub_industry.
    """
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    try:
        tables = pd.read_html(url)
        if not tables:
            print("No tables found on Wikipedia S&P 500 page.")
            return pd.DataFrame()

        df = tables[0]

        # Standardize column names
        col_map = {}
        for col in df.columns:
            c = str(col).lower()
            if "symbol" in c or "ticker" in c:
                col_map[col] = "symbol"
            elif "security" in c or "company" in c:
                col_map[col] = "company"
            elif "sector" in c and "sub" not in c:
                col_map[col] = "sector"
            elif "sub" in c and "industry" in c:
                col_map[col] = "sub_industry"

        df = df.rename(columns=col_map)

        # Keep only relevant columns
        keep_cols = [c for c in ["symbol", "company", "sector", "sub_industry"]
                     if c in df.columns]
        df = df[keep_cols].copy()

        if "symbol" in df.columns:
            df["symbol"] = df["symbol"].astype(str).str.strip()
            # Fix tickers with dots (BRK.B → BRK-B for yfinance)
            df["symbol_yf"] = df["symbol"].str.replace(".", "-", regex=False)

        print(f"  Fetched {len(df)} S&P 500 components.")
        return df

    except Exception as e:
        print(f"Error fetching S&P 500 components: {e}")
        return pd.DataFrame()


# =====================================================================
# TWSE Components
# =====================================================================


def fetch_twse_components():
    """Fetch Taiwan Stock Exchange weighted index components.

    Uses TWSE open data for listed companies with P/E ratios.

    Returns:
        pd.DataFrame with columns: symbol, company, pe_ratio, dividend_yield.
    """
    url = "https://www.twse.com.tw/exchangeReport/BWIBBU_d?response=json&selectType=ALL"
    try:
        import urllib.request
        import json

        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if "data" not in data:
            print("No data in TWSE response.")
            return pd.DataFrame()

        fields = data.get("fields", [])
        rows = data["data"]

        df = pd.DataFrame(rows, columns=fields if fields else None)

        # TWSE fields are typically:
        # 證券代號, 證券名稱, 殖利率(%), 股利年度, 本益比, 股價淨值比, 財報年/季
        if len(df.columns) >= 5:
            df = df.rename(columns={
                df.columns[0]: "symbol",
                df.columns[1]: "company",
                df.columns[2]: "dividend_yield",
                df.columns[4]: "pe_ratio",
            })

            df["symbol"] = df["symbol"].astype(str).str.strip()
            df["company"] = df["company"].astype(str).str.strip()
            df["pe_ratio"] = pd.to_numeric(df["pe_ratio"], errors="coerce")
            df["dividend_yield"] = pd.to_numeric(df["dividend_yield"], errors="coerce")

            # Add .TW suffix for yfinance
            df["symbol_yf"] = df["symbol"] + ".TW"

        print(f"  Fetched {len(df)} TWSE components.")
        return df

    except Exception as e:
        print(f"Error fetching TWSE components: {e}")
        # Fallback: return empty DataFrame
        return pd.DataFrame()


# =====================================================================
# Sorting and Enrichment
# =====================================================================


def sort_by_pe(components_df, ascending=True):
    """Sort components by P/E ratio.

    Args:
        components_df: DataFrame with 'pe_ratio' column.
        ascending: Sort order (True=lowest P/E first).

    Returns:
        Sorted DataFrame.
    """
    if components_df.empty or "pe_ratio" not in components_df.columns:
        return components_df

    df = components_df.copy()
    df = df[df["pe_ratio"].notna() & (df["pe_ratio"] > 0)]
    df = df.sort_values("pe_ratio", ascending=ascending).reset_index(drop=True)
    return df


def enrich_with_metrics(components_df, max_stocks=50):
    """Add P/E, dividend yield, market cap via yfinance batch fetch.

    Args:
        components_df: DataFrame with 'symbol' (or 'symbol_yf') column.
        max_stocks: Maximum number of stocks to enrich (to limit API calls).

    Returns:
        DataFrame with additional columns: pe_ratio, dividend_yield, market_cap.
    """
    if components_df.empty:
        return components_df

    df = components_df.head(max_stocks).copy()

    # Use symbol_yf if available, else symbol
    sym_col = "symbol_yf" if "symbol_yf" in df.columns else "symbol"

    # Initialize columns if not present
    for col in ["pe_ratio", "dividend_yield", "market_cap", "current_price"]:
        if col not in df.columns:
            df[col] = np.nan

    symbols = df[sym_col].tolist()
    print(f"  Enriching {len(symbols)} stocks with yfinance data...")

    # Batch fetch using yfinance download for prices
    batch_size = 20
    for i in range(0, len(symbols), batch_size):
        batch = symbols[i:i + batch_size]
        batch_str = " ".join(batch)
        print(f"    Batch {i // batch_size + 1}: {len(batch)} stocks...")

        for sym in batch:
            try:
                t = yf.Ticker(sym)
                info = t.info
                idx = df[df[sym_col] == sym].index
                if len(idx) == 0:
                    continue
                idx = idx[0]

                pe = info.get("trailingPE")
                if pe is not None:
                    df.at[idx, "pe_ratio"] = pe

                dy = info.get("dividendYield")
                if dy is not None:
                    df.at[idx, "dividend_yield"] = dy

                mc = info.get("marketCap")
                if mc is not None:
                    df.at[idx, "market_cap"] = mc

                price = info.get("currentPrice") or info.get("regularMarketPrice")
                if price is not None:
                    df.at[idx, "current_price"] = price

            except Exception:
                continue

    return df


def export_components(components_df, output_path="index_components.xlsx",
                      index_name="S&P 500"):
    """Write index components to Excel.

    Args:
        components_df: DataFrame with component data.
        output_path: Output file path.
        index_name: Name of the index for the header.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = index_name[:31]  # Excel sheet name max 31 chars

    ws.cell(row=1, column=1, value=f"{index_name} Components")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.today().strftime('%Y-%m-%d')}")
    ws.cell(row=2, column=3, value=f"Total: {len(components_df)}")

    # Write headers
    display_cols = [c for c in ["symbol", "company", "sector", "pe_ratio",
                                "dividend_yield", "market_cap", "current_price"]
                    if c in components_df.columns]
    for ci, col in enumerate(display_cols, start=1):
        ws.cell(row=4, column=ci, value=col)

    # Write data
    for ri, (_, row) in enumerate(components_df.iterrows(), start=5):
        for ci, col in enumerate(display_cols, start=1):
            val = row.get(col)
            ws.cell(row=ri, column=ci, value=_safe_val(val))

    wb.save(output_path)
    print(f"Saved: {output_path}")


# =====================================================================
# CLI Entry Point
# =====================================================================

def main():
    index_type = sys.argv[1].lower() if len(sys.argv) > 1 else "sp500"
    enrich = "--enrich" in sys.argv
    top_n = 50

    # Parse --top N
    for i, arg in enumerate(sys.argv):
        if arg == "--top" and i + 1 < len(sys.argv):
            try:
                top_n = int(sys.argv[i + 1])
            except ValueError:
                pass

    if index_type in ("sp500", "sp", "s&p"):
        print(f"{'='*50}")
        print("  Fetching S&P 500 Components")
        print(f"{'='*50}")

        df = fetch_sp500_components()
        if df.empty:
            print("Failed to fetch S&P 500 components.")
            return

        if enrich:
            print("\nEnriching with financial metrics...")
            df = enrich_with_metrics(df, max_stocks=top_n)

        if "pe_ratio" in df.columns and df["pe_ratio"].notna().any():
            df_sorted = sort_by_pe(df)
            print(f"\nTop 20 by lowest P/E:")
            for _, row in df_sorted.head(20).iterrows():
                sym = row.get("symbol", "")
                pe = row.get("pe_ratio", "")
                name = row.get("company", "")[:30]
                pe_str = f"{pe:.1f}" if isinstance(pe, (int, float)) and not np.isnan(pe) else "N/A"
                print(f"  {sym:8s} PE={pe_str:>8s}  {name}")

        output = "sp500_components.xlsx"
        export_components(df, output, "S&P 500")

    elif index_type in ("twse", "tw", "taiwan"):
        print(f"{'='*50}")
        print("  Fetching TWSE Components")
        print(f"{'='*50}")

        df = fetch_twse_components()
        if df.empty:
            print("Failed to fetch TWSE components.")
            return

        if enrich:
            print("\nEnriching with financial metrics...")
            df = enrich_with_metrics(df, max_stocks=top_n)

        if "pe_ratio" in df.columns and df["pe_ratio"].notna().any():
            df_sorted = sort_by_pe(df)
            print(f"\nTop 20 by lowest P/E:")
            for _, row in df_sorted.head(20).iterrows():
                sym = row.get("symbol", "")
                pe = row.get("pe_ratio", "")
                name = row.get("company", "")[:20]
                pe_str = f"{pe:.1f}" if isinstance(pe, (int, float)) and not np.isnan(pe) else "N/A"
                print(f"  {sym:8s} PE={pe_str:>8s}  {name}")

        output = "twse_components.xlsx"
        export_components(df, output, "TWSE")

    else:
        print(f"Unknown index type: {index_type}")
        print("Usage: python index_components.py [sp500|twse] [--enrich] [--top N]")


if __name__ == "__main__":
    main()
