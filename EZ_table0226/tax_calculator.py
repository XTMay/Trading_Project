"""
tax_calculator.py
Taiwan Dividend Tax Engine for 盈再表.

Implements VBA Module9 logic for Taiwan stock dividend tax calculations:
  - Dividend tax withholding (扣繳稅款)
  - KY stock identification (KY 股判定)
  - Ex-dividend matching (除息配對)
  - Annual tax summary (年度稅務彙總)

Taiwan dividend tax rules:
  - Dividends < 20,000 NTD: (dividend + supplement) × rate - 10
  - Dividends >= 20,000 NTD: dividend × rate × (1 - supplement_rate) + supplement × rate - 10
  - KY stocks have different supplement fee rate (2.11%)

Usage:
    python tax_calculator.py                       # demo with examples
    python tax_calculator.py holdings.xlsx 2024     # compute tax for year
"""

import sys
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook

from generate_report_summary import _safe_val

# =====================================================================
# Constants (from VBA Module9: [i10], [i11], [i12])
# =====================================================================

DIVIDEND_THRESHOLD = 20000   # NTD threshold for different tax treatment
SUPPLEMENT_RATE_KY = 0.0211  # 2.11% supplementary fee for KY stocks
FLAT_FEE = 10               # NTD flat filing fee
DEFAULT_TAX_RATE = 0.25      # Default withholding tax rate (25%)


# =====================================================================
# Core Tax Calculations
# =====================================================================


def calc_dividend_tax(dividend_per_share, shares, is_ky=False,
                      threshold=DIVIDEND_THRESHOLD,
                      tax_rate=DEFAULT_TAX_RATE,
                      supplement_rate=SUPPLEMENT_RATE_KY):
    """Compute dividend tax withholding for a single stock.

    From VBA Module9:
      If total_dividend < threshold:
        tax = (dividend_base + supplement) × rate - flat_fee
      If total_dividend >= threshold:
        tax = dividend_base × rate × (1 - supplement_rate) + supplement × rate - flat_fee

    Args:
        dividend_per_share: Dividend amount per share (NTD).
        shares: Number of shares held.
        is_ky: Whether the stock is a KY stock.
        threshold: Dividend threshold (default 20,000 NTD).
        tax_rate: Withholding tax rate (default 0.25).
        supplement_rate: Supplementary fee rate for KY stocks.

    Returns:
        dict with tax calculation details.
    """
    if dividend_per_share is None or shares is None:
        return {"total_dividend": 0, "tax": 0, "net_dividend": 0}

    total_dividend = dividend_per_share * shares

    if total_dividend <= 0:
        return {"total_dividend": 0, "tax": 0, "net_dividend": 0}

    if total_dividend < threshold:
        # Simple formula: (dividend + supplement) × rate - flat_fee
        tax = total_dividend * tax_rate - FLAT_FEE
    else:
        if is_ky:
            # KY stock: dividend × rate × (1 - supplement_rate) + supplement × rate - flat_fee
            tax = (total_dividend * tax_rate * (1 - supplement_rate)
                   + total_dividend * supplement_rate * tax_rate
                   - FLAT_FEE)
        else:
            # Non-KY: standard withholding
            tax = total_dividend * tax_rate - FLAT_FEE

    tax = max(tax, 0)  # Tax cannot be negative
    net_dividend = total_dividend - tax

    return {
        "total_dividend": round(total_dividend, 2),
        "tax": round(tax, 2),
        "net_dividend": round(net_dividend, 2),
        "tax_rate": tax_rate,
        "is_ky": is_ky,
        "effective_rate": round(tax / total_dividend, 4) if total_dividend > 0 else 0,
    }


def is_ky_stock(stock_code):
    """Check if stock code is a KY stock (foreign company listed in Taiwan).

    KY stocks are identified by "KY" suffix in the stock name or code.
    Common pattern: 股票代号 contains "KY" or stock name contains "KY".

    Args:
        stock_code: Stock code string (e.g., "6547-KY", "KY2762").

    Returns:
        bool: True if KY stock.
    """
    if not stock_code:
        return False
    code = str(stock_code).upper().strip()
    return "KY" in code


def match_ex_dividends(holdings, year=None):
    """Match holdings with ex-dividend dates and amounts.

    Fetches dividend data from yfinance for each holding.

    Args:
        holdings: DataFrame with 'code' and 'shares' columns.
        year: Target year for dividend matching (default: current year).

    Returns:
        DataFrame with additional columns:
            ex_date, dividend_per_share, total_dividend.
    """
    if holdings is None or holdings.empty:
        return holdings

    if year is None:
        year = datetime.today().year

    df = holdings.copy()
    df["ex_date"] = None
    df["dividend_per_share"] = 0.0
    df["total_dividend"] = 0.0

    for idx, row in df.iterrows():
        code = row["code"]
        shares = row.get("shares", 0)

        try:
            ticker = yf.Ticker(code)
            actions = ticker.actions
            if actions is None or actions.empty:
                continue

            # Filter dividends for the target year
            divs = actions[actions["Dividends"] > 0]
            divs = divs[divs.index.year == year]

            if divs.empty:
                continue

            # Sum all dividends in the year for this stock
            total_div_per_share = divs["Dividends"].sum()
            last_ex_date = divs.index[-1]

            df.at[idx, "ex_date"] = last_ex_date.strftime("%Y-%m-%d")
            df.at[idx, "dividend_per_share"] = total_div_per_share
            df.at[idx, "total_dividend"] = total_div_per_share * shares

        except Exception as e:
            print(f"  Warning: Could not fetch dividends for {code}: {e}")

    return df


def generate_tax_summary(year, holdings_with_dividends):
    """Generate annual tax report by stock.

    Args:
        year: Tax year.
        holdings_with_dividends: DataFrame from match_ex_dividends()
            with columns: code, shares, dividend_per_share, total_dividend.

    Returns:
        dict with summary and per-stock tax details.
    """
    if holdings_with_dividends is None or holdings_with_dividends.empty:
        return {"year": year, "total_tax": 0, "total_net": 0, "stocks": []}

    df = holdings_with_dividends.copy()
    stocks = []
    total_tax = 0
    total_gross = 0
    total_net = 0
    ky_tax = 0
    non_ky_tax = 0

    for _, row in df.iterrows():
        code = row["code"]
        shares = row.get("shares", 0)
        div_ps = row.get("dividend_per_share", 0)

        if div_ps <= 0:
            continue

        ky = is_ky_stock(code)
        tax_result = calc_dividend_tax(div_ps, shares, is_ky=ky)

        stock_info = {
            "code": code,
            "shares": shares,
            "is_ky": ky,
            "ex_date": row.get("ex_date", ""),
            "dividend_per_share": div_ps,
            **tax_result,
        }
        stocks.append(stock_info)

        total_tax += tax_result["tax"]
        total_gross += tax_result["total_dividend"]
        total_net += tax_result["net_dividend"]

        if ky:
            ky_tax += tax_result["tax"]
        else:
            non_ky_tax += tax_result["tax"]

    return {
        "year": year,
        "total_gross_dividend": round(total_gross, 2),
        "total_tax": round(total_tax, 2),
        "total_net_dividend": round(total_net, 2),
        "ky_stock_tax": round(ky_tax, 2),
        "non_ky_stock_tax": round(non_ky_tax, 2),
        "effective_rate": round(total_tax / total_gross, 4) if total_gross > 0 else 0,
        "stocks": stocks,
    }


def export_tax_report(tax_summary, output_path="tax_report.xlsx"):
    """Write tax report to Excel.

    Args:
        tax_summary: dict from generate_tax_summary().
        output_path: Output file path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Summary"

    year = tax_summary.get("year", "")
    ws.cell(row=1, column=1, value=f"{year} Dividend Tax Report (股利稅務報告)")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.today().strftime('%Y-%m-%d')}")

    # Summary section
    row = 4
    items = [
        ("Total Gross Dividend (股利總額)", tax_summary.get("total_gross_dividend")),
        ("Total Tax (扣繳稅款)", tax_summary.get("total_tax")),
        ("Total Net Dividend (淨股利)", tax_summary.get("total_net_dividend")),
        ("KY Stock Tax (KY股稅款)", tax_summary.get("ky_stock_tax")),
        ("Non-KY Stock Tax (非KY股稅款)", tax_summary.get("non_ky_stock_tax")),
        ("Effective Tax Rate (有效稅率)", tax_summary.get("effective_rate")),
    ]
    for label, val in items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=_safe_val(val))
        row += 1

    # Per-stock detail
    row += 1
    ws.cell(row=row, column=1, value="Per-Stock Detail")
    row += 1
    headers = [
        "Code", "KY?", "Shares", "Ex-Date", "Div/Share",
        "Total Dividend", "Tax", "Net Dividend", "Effective Rate",
    ]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=row, column=ci, value=h)

    for stock in tax_summary.get("stocks", []):
        row += 1
        ws.cell(row=row, column=1, value=stock.get("code", ""))
        ws.cell(row=row, column=2, value="Yes" if stock.get("is_ky") else "No")
        ws.cell(row=row, column=3, value=_safe_val(stock.get("shares")))
        ws.cell(row=row, column=4, value=stock.get("ex_date", ""))
        ws.cell(row=row, column=5, value=_safe_val(stock.get("dividend_per_share")))
        ws.cell(row=row, column=6, value=_safe_val(stock.get("total_dividend")))
        ws.cell(row=row, column=7, value=_safe_val(stock.get("tax")))
        ws.cell(row=row, column=8, value=_safe_val(stock.get("net_dividend")))
        ws.cell(row=row, column=9, value=_safe_val(stock.get("effective_rate")))

    wb.save(output_path)
    print(f"Saved: {output_path}")


# =====================================================================
# CLI Entry Point
# =====================================================================

def main():
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
        year = int(sys.argv[2]) if len(sys.argv) > 2 else datetime.today().year

        print(f"Loading holdings from: {excel_path}")
        print(f"Tax year: {year}")

        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            print(f"Error reading {excel_path}: {e}")
            return

        # Flexible column mapping
        col_map = {}
        for col in df.columns:
            c = str(col).lower().strip()
            if c in ("code", "ticker", "symbol", "股票代号"):
                col_map[col] = "code"
            elif c in ("shares", "股数", "持股"):
                col_map[col] = "shares"

        df = df.rename(columns=col_map)

        if "code" not in df.columns or "shares" not in df.columns:
            print("Missing required columns: code, shares")
            return

        print("Fetching dividend data...")
        df = match_ex_dividends(df, year)

        print("Computing tax...")
        summary = generate_tax_summary(year, df)

        print(f"\n{'='*50}")
        print(f"  {year} Dividend Tax Summary")
        print(f"{'='*50}")
        print(f"  Gross Dividends: {summary['total_gross_dividend']:,.2f}")
        print(f"  Total Tax:       {summary['total_tax']:,.2f}")
        print(f"  Net Dividends:   {summary['total_net_dividend']:,.2f}")
        if summary['effective_rate'] > 0:
            print(f"  Effective Rate:  {summary['effective_rate']:.2%}")

        if summary["stocks"]:
            print(f"\n  Per-Stock Detail:")
            for s in summary["stocks"]:
                ky_tag = " [KY]" if s["is_ky"] else ""
                print(f"    {s['code']}{ky_tag}: "
                      f"div={s['total_dividend']:,.0f}, "
                      f"tax={s['tax']:,.0f}, "
                      f"net={s['net_dividend']:,.0f}")

        output = f"tax_report_{year}.xlsx"
        export_tax_report(summary, output)

    else:
        # Demo mode
        print("Taiwan Dividend Tax Calculator - Demo")
        print("=" * 50)

        # Example 1: Non-KY stock, small dividend
        print("\nExample 1: Non-KY stock, dividend < 20,000 NTD")
        result1 = calc_dividend_tax(5.0, 1000, is_ky=False)
        print(f"  1000 shares × $5.00 = ${result1['total_dividend']:,.2f}")
        print(f"  Tax: ${result1['tax']:,.2f}")
        print(f"  Net: ${result1['net_dividend']:,.2f}")
        print(f"  Effective rate: {result1['effective_rate']:.2%}")

        # Example 2: Non-KY stock, large dividend
        print("\nExample 2: Non-KY stock, dividend >= 20,000 NTD")
        result2 = calc_dividend_tax(10.0, 5000, is_ky=False)
        print(f"  5000 shares × $10.00 = ${result2['total_dividend']:,.2f}")
        print(f"  Tax: ${result2['tax']:,.2f}")
        print(f"  Net: ${result2['net_dividend']:,.2f}")
        print(f"  Effective rate: {result2['effective_rate']:.2%}")

        # Example 3: KY stock
        print("\nExample 3: KY stock, dividend >= 20,000 NTD")
        result3 = calc_dividend_tax(8.0, 5000, is_ky=True)
        print(f"  5000 shares × $8.00 = ${result3['total_dividend']:,.2f}")
        print(f"  Tax: ${result3['tax']:,.2f}")
        print(f"  Net: ${result3['net_dividend']:,.2f}")
        print(f"  Effective rate: {result3['effective_rate']:.2%}")

        # KY detection
        print("\nKY Stock Detection:")
        test_codes = ["2330", "6547-KY", "KY2762", "2454", "91APP-KY"]
        for code in test_codes:
            print(f"  {code}: {'KY' if is_ky_stock(code) else 'Non-KY'}")


if __name__ == "__main__":
    main()
